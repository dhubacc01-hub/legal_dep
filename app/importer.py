from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any


HEADER_ALIASES = {
    "entry_date": ["дата внесения", "дата создания", "дата добавления", "дата"],
    "contract_date": ["дата контракта", "дата договора"],
    "category": ["категория", "статус", "статус дела"],
    "client_name": ["фио клиента", "фио", "клиент", "ф.и.о. клиента"],
    "contract_number": ["№ договора", "номер договора", "договор", "номер контракта"],
    "last_missed_payment_date": [
        "дата последнего неисполненного платежа",
        "дата последнего платежа",
        "дата просрочки",
    ],
    "company": ["компания", "бренд", "юр лицо"],
    "city": ["город", "населенный пункт"],
    "court": ["суд", "текущий суд"],
    "claim_sent": ["претензия", "отправлена претензия"],
    "claim_sent_date": ["дата отправки претензии"],
    "debt_amount": ["сумма долга", "долг", "сумма долга (тг)"],
    "lawsuit_sent": ["направлен иск", "иск подан", "отправлен иск", "направлено заявление"],
    "lawsuit_sent_date": ["дата отправки иска", "дата подачи иска", "дата отправки"],
    "lawsuit_accepted": ["иск принят", "принят иск"],
    "hearing_date": ["дата заседания", "заседание"],
    "decision_exists": ["есть решение", "решение есть", "вынесено решение"],
    "decision": ["решение", "итог"],
    "decision_payout": ["сумма выплаты по решению", "выплата по решению"],
    "received_amount": ["получено", "получено (тг)", "оплачено", "поступило"],
    "comment": ["комментарий", "коммент", "примечание"],
    "case_number": ["номер дела", "номер дела и суд", "№ дела", "номер дела и суд"],
    "imported_claim_sent_days": ["кол дней с отправки претензии", "кол-во дней с отправки претензии"],
    "imported_debt_days": ["количество дней долга", "кол-во дней долга"],
    "imported_penalty_amount": ["пеня (тг)", "пеня"],
    "imported_state_duty_amount": ["сумма гос пошлины (тг)", "сумма гос. пошлины (тг)"],
    "imported_total_amount": ["общая сумма (тг)", "общая сумма"],
}


def load_rows_from_path(path_value: str) -> list[dict[str, Any]]:
    path = Path(path_value).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")
    if not path.is_file():
        raise IsADirectoryError(f"Ожидался файл, а не папка: {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return load_rows_from_xlsx(path)
    if suffix == ".csv":
        return load_rows_from_csv(path)
    raise ValueError("Поддерживаются только .xlsx, .xlsm и .csv файлы.")


def load_rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return []

    headers = [normalize_header_cell(value) for value in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for row_values in rows[1:]:
        if row_is_empty(row_values):
            continue
        raw_row = {
            headers[index] or f"column_{index + 1}": row_values[index] if index < len(row_values) else None
            for index in range(len(headers))
        }
        data_rows.append(raw_row)
    return data_rows


def load_rows_from_csv(path: Path) -> list[dict[str, Any]]:
    raw_bytes = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Не удалось определить кодировку CSV файла.")

    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, Any]] = []
    for row in reader:
        if row_is_empty(row.values()):
            continue
        rows.append({normalize_header_cell(key): value for key, value in row.items()})
    return rows


def normalize_import_row(
    raw_row: dict[str, Any],
    *,
    categories: list[str],
    decisions: list[str],
    companies: list[str],
    cities: list[str],
    courts: list[str],
) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    warnings: list[str] = []
    errors: list[str] = []

    headers_map = build_headers_map(raw_row)

    for canonical_field in HEADER_ALIASES:
        raw_value = find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES[canonical_field])
        if canonical_field in {"entry_date", "contract_date"}:
            normalized[canonical_field] = parse_date_cell(raw_value)
        elif canonical_field in {
            "last_missed_payment_date",
            "claim_sent_date",
            "lawsuit_sent_date",
            "hearing_date",
        }:
            normalized[canonical_field] = parse_date_cell(raw_value)
            if raw_value not in (None, "") and normalized[canonical_field] is None:
                errors.append(f"Не удалось распознать дату в поле «{canonical_field}».")
        elif canonical_field in {
            "claim_sent",
            "lawsuit_sent",
            "lawsuit_accepted",
            "decision_exists",
        }:
            normalized[canonical_field] = parse_bool_cell(raw_value)
            if raw_value not in (None, "") and normalized[canonical_field] is None:
                warnings.append(f"Не удалось однозначно распознать Да/Нет в поле «{canonical_field}».")
        elif canonical_field in {
            "debt_amount",
            "decision_payout",
            "received_amount",
            "imported_claim_sent_days",
            "imported_debt_days",
            "imported_penalty_amount",
            "imported_state_duty_amount",
            "imported_total_amount",
        }:
            normalized[canonical_field] = parse_number_cell(raw_value)
            if raw_value not in (None, "") and normalized[canonical_field] is None:
                errors.append(f"Не удалось распознать число в поле «{canonical_field}».")
        else:
            normalized[canonical_field] = clean_text(raw_value)

    normalized["category"] = normalize_lookup(normalized.get("category"), categories)
    normalized["decision"] = normalize_lookup(normalized.get("decision"), decisions)
    normalized["company"] = normalize_lookup(normalized.get("company"), companies)
    normalized["city"] = normalize_lookup(normalized.get("city"), cities)
    normalized["court"] = normalize_lookup(normalized.get("court"), courts)

    detect_unknown_reference(
        raw_row,
        headers_map,
        warnings,
        field="category",
        aliases=HEADER_ALIASES["category"],
        normalized_value=normalized.get("category"),
    )
    detect_unknown_reference(
        raw_row,
        headers_map,
        warnings,
        field="decision",
        aliases=HEADER_ALIASES["decision"],
        normalized_value=normalized.get("decision"),
    )
    detect_unknown_reference(
        raw_row,
        headers_map,
        warnings,
        field="company",
        aliases=HEADER_ALIASES["company"],
        normalized_value=normalized.get("company"),
    )
    detect_unknown_reference(
        raw_row,
        headers_map,
        warnings,
        field="city",
        aliases=HEADER_ALIASES["city"],
        normalized_value=normalized.get("city"),
    )
    detect_unknown_reference(
        raw_row,
        headers_map,
        warnings,
        field="court",
        aliases=HEADER_ALIASES["court"],
        normalized_value=normalized.get("court"),
    )

    required_text_fields = ["client_name", "contract_number", "company", "city", "court"]
    for field_name in required_text_fields:
        if not normalized.get(field_name):
            errors.append(f"Обязательное поле «{field_name}» пустое или не найдено.")

    if normalized.get("last_missed_payment_date") is None:
        errors.append("Не найдена дата последнего неисполненного платежа.")

    if normalized.get("debt_amount") is None:
        errors.append("Не найдена сумма долга.")

    return {
        "normalized_data": normalized,
        "warnings": dedupe_messages(warnings),
        "errors": dedupe_messages(errors),
    }


def normalize_header_cell(value: Any) -> str:
    text = clean_text(value)
    return text or ""


def row_is_empty(values: Any) -> bool:
    return all(clean_text(value) in (None, "") for value in values)


def build_headers_map(raw_row: dict[str, Any]) -> dict[str, str]:
    return {normalize_header_key(header): header for header in raw_row}


def find_value_by_aliases(raw_row: dict[str, Any], headers_map: dict[str, str], aliases: list[str]) -> Any:
    for alias in aliases:
        matched_header = headers_map.get(normalize_header_key(alias))
        if matched_header is not None:
            return raw_row.get(matched_header)
    return None


def detect_unknown_reference(
    raw_row: dict[str, Any],
    headers_map: dict[str, str],
    warnings: list[str],
    *,
    field: str,
    aliases: list[str],
    normalized_value: Any,
) -> None:
    raw_value = find_value_by_aliases(raw_row, headers_map, aliases)
    raw_text = clean_text(raw_value)
    if raw_text and not normalized_value:
        warnings.append(f"Поле «{field}» содержит неизвестное значение: {raw_text}.")


def parse_bool_cell(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value

    text = clean_text(value)
    if text is None:
        return None

    normalized = normalize_header_key(text)
    truthy = {"да", "true", "1", "есть", "y", "yes", "подан", "принят"}
    falsy = {"нет", "false", "0", "no", "n", "неподан", "непринят"}
    if normalized in truthy:
        return True
    if normalized in falsy:
        return False
    return None


def parse_date_cell(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = clean_text(value)
    if not text:
        return None

    for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%y"):
        try:
            parsed = datetime.strptime(text, pattern)
            if parsed.year < 100:
                parsed = parsed.replace(year=2000 + parsed.year)
            return parsed.date().isoformat()
        except ValueError:
            continue
    return None


def parse_number_cell(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)

    text = clean_text(value)
    if not text:
        return None

    normalized = text.replace("\u00a0", "").replace(" ", "")
    normalized = normalized.replace("тг", "").replace("₸", "")
    if normalized.count(",") == 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", ".")
    elif normalized.count(",") > 0 and normalized.count(".") > 0:
        normalized = normalized.replace(",", "")

    normalized = re.sub(r"[^0-9.\-]", "", normalized)
    if not normalized or normalized in {"-", ".", "-."}:
        return None

    try:
        return round(float(normalized), 2)
    except ValueError:
        return None


def normalize_lookup(value: str | None, options: list[str]) -> str | None:
    text = clean_text(value)
    if not text:
        return None

    normalized_source = normalize_header_key(text)
    for option in options:
        if normalize_header_key(option) == normalized_source:
            return option
    return None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_header_key(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", value).replace("ё", "е").replace("Ё", "Е")
    return re.sub(r"[^a-zA-Zа-яА-Я0-9]+", "", ascii_text).lower()


def dedupe_messages(items: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result
