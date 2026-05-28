from __future__ import annotations

import json
import re
from calendar import monthrange
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Any

from fastapi import Depends, FastAPI, HTTPException, Request, Response, status
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

from app.auth_service import (
    PASSWORD_MIN_LENGTH,
    ROLE_ADMIN,
    ROLE_LAWYER,
    ROLE_OWNER,
    SESSION_COOKIE_NAME,
    USER_ROLES,
    build_session_expiry,
    create_session_token,
    ensure_bootstrap_users,
    hash_password,
    serialize_user,
    utc_now,
    verify_password,
)
from app.court_catalog import merge_court_catalog
from app.database import get_connection, init_db
from app.disell_api import DiSellApiClient, DiSellApiError
from app.importer import (
    HEADER_ALIASES,
    build_headers_map,
    find_value_by_aliases,
    load_rows_from_path,
    normalize_import_row,
)
from app.reference_data import (
    CATEGORIES,
    COMPANIES,
    COUNTRY_LABELS,
    COURTS_BY_CITY,
    DECISIONS,
    DEFAULT_COUNTRY,
    KAZAKHSTAN_CITIES,
    SUPPORTED_COUNTRIES,
    UZBEKISTAN_COURT_CATALOG,
    get_companies_by_country,
)
from app.schemas import (
    ClaimPdfGenerateRequest,
    CrmDebtorLookupResponse,
    CourtCreate,
    DebtorCreate,
    DebtorUpdate,
    AuthMeResponse,
    ChangePasswordRequest,
    ImportApplyRequest,
    ImportPreviewRequest,
    LawsuitPdfGenerateRequest,
    LoginRequest,
    UserCreateRequest,
    UserView,
)

BASE_DIR = Path(__file__).resolve().parent
PROJECTS_DIR = BASE_DIR.parent.parent
GENERATED_DIR = BASE_DIR.parent / "data" / "generated"
REQUISITES_PATH_CANDIDATES = [
    PROJECTS_DIR / "реквизиты.xlsx",
    PROJECTS_DIR / "Реквизиты.xlsx",
]
PDF_BROWSER_CANDIDATES = [
    Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
    Path(r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
    Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
]
SERIF_FONT_REGULAR = Path(r"C:\Windows\Fonts\times.ttf")
SERIF_FONT_BOLD = Path(r"C:\Windows\Fonts\timesbd.ttf")

DEFAULT_CATEGORY = "Новый"
CATEGORY_PREPARE_LAWSUIT = "Готовим иск"
CATEGORY_LAWSUIT_FILED = "Иск подан"
CATEGORY_LAWSUIT_CLOSED = "Иск закрыт"
CATEGORY_NO_JURISDICTION = "Неподсудно"
CATEGORY_LIMITATION_EXPIRED = "Прошел срок исковой давности"
CATEGORY_SMALL_DEBT = "Маленькая сумма долга"
CATEGORY_WAITING_FOR_CLAIM_RESPONSE = "Ожидаем ответа по претензии"
CATEGORY_RETURNED_TO_LEGAL = "Возврат в работу Юр. Отдела"
CATEGORY_DEBT_CLOSED = "Долг закрыт"

DECISION_SATISFY = "Удовлетворить"
DECISION_PARTIAL = "Частично"
DECISION_SETTLEMENT = "По соглашению сторон"
DECISION_REFUSAL = "Отказ в иске"
DECISION_RETURN = "Возврат иска"

PRIORITY_CATEGORY_OVERRIDES = {
    "Закрытая компания",
    "Не должник",
    "Клиент частично оплачивает",
    "Требуется проверка решения в кабинете",
    "Передать на ЧСИ",
}

DECISIONS_THAT_CLOSE_LAWSUIT = {
    DECISION_SATISFY,
    DECISION_PARTIAL,
    DECISION_SETTLEMENT,
}

SAFE_IMPORT_DECISION_CATEGORY_MAP = {
    "Передать на ЧСИ": "Передать на ЧСИ",
    "Прошел срок исковой давности": CATEGORY_LIMITATION_EXPIRED,
    "Маленькая сумма долга": CATEGORY_SMALL_DEBT,
    "Долг закрыт!": CATEGORY_DEBT_CLOSED,
    "Закрытая компания": "Закрытая компания",
    "Не должник": "Не должник",
    "Оплата по претензии": "Оплата по претензии",
}

SAFE_IMPORT_STAGE2_RULES: dict[tuple[str, str], dict[str, Any]] = {
    (CATEGORY_LAWSUIT_CLOSED, DECISION_SETTLEMENT): {
        "category": CATEGORY_LAWSUIT_CLOSED,
        "decision": DECISION_SETTLEMENT,
        "decision_exists": True,
        "claim_sent": True,
        "lawsuit_sent": True,
        "lawsuit_accepted": True,
        "imported_category_override": False,
    },
    (CATEGORY_LAWSUIT_CLOSED, DECISION_SATISFY): {
        "category": CATEGORY_LAWSUIT_CLOSED,
        "decision": DECISION_SATISFY,
        "decision_exists": True,
        "claim_sent": True,
        "lawsuit_sent": True,
        "lawsuit_accepted": True,
        "imported_category_override": False,
    },
    (CATEGORY_LAWSUIT_CLOSED, DECISION_PARTIAL): {
        "category": CATEGORY_LAWSUIT_CLOSED,
        "decision": DECISION_PARTIAL,
        "decision_exists": True,
        "claim_sent": True,
        "lawsuit_sent": True,
        "lawsuit_accepted": True,
        "imported_category_override": False,
    },
    ("Проверка решения в кабинете", ""): {
        "category": "Требуется проверка решения в кабинете",
        "decision": None,
        "decision_exists": False,
        "claim_sent": True,
        "lawsuit_sent": True,
        "lawsuit_accepted": True,
        "imported_category_override": True,
    },
    (CATEGORY_LAWSUIT_FILED, ""): {
        "category": CATEGORY_LAWSUIT_FILED,
        "decision": None,
        "decision_exists": False,
        "claim_sent": True,
        "lawsuit_sent": True,
        "lawsuit_accepted": False,
        "imported_category_override": False,
    },
    ("Возврат иска", "Отказ"): {
        "category": CATEGORY_NO_JURISDICTION,
        "decision": DECISION_REFUSAL,
        "decision_exists": True,
        "claim_sent": True,
        "lawsuit_sent": True,
        "lawsuit_accepted": True,
        "imported_category_override": False,
    },
    ("Возврат иска", DECISION_RETURN): {
        "category": CATEGORY_RETURNED_TO_LEGAL,
        "decision": DECISION_RETURN,
        "decision_exists": True,
        "claim_sent": True,
        "lawsuit_sent": True,
        "lawsuit_accepted": True,
        "imported_category_override": False,
    },
}

app = FastAPI(title="Legal Department", version="0.3.0")
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
UZ_BRV_AMOUNT = 412000.0


@app.on_event("startup")
def on_startup() -> None:
    init_db()
    with get_connection() as connection:
        ensure_bootstrap_users(connection)
        connection.commit()


def get_session_user(connection, session_token: str | None) -> dict[str, Any] | None:
    if not session_token:
        return None

    row = connection.execute(
        """
        SELECT u.*
        FROM user_sessions s
        JOIN users u ON u.id = s.user_id
        WHERE s.session_token = ?
        """,
        (session_token,),
    ).fetchone()
    if row is None:
        return None

    expires_at_row = connection.execute(
        "SELECT expires_at FROM user_sessions WHERE session_token = ?",
        (session_token,),
    ).fetchone()
    expires_at_value = expires_at_row["expires_at"] if expires_at_row else None
    if not expires_at_value:
        connection.execute("DELETE FROM user_sessions WHERE session_token = ?", (session_token,))
        connection.commit()
        return None

    expires_at = datetime.fromisoformat(str(expires_at_value))
    if expires_at <= utc_now():
        connection.execute("DELETE FROM user_sessions WHERE session_token = ?", (session_token,))
        connection.commit()
        return None

    user = dict(row)
    if not parse_bool(user.get("is_active")):
        connection.execute("DELETE FROM user_sessions WHERE session_token = ?", (session_token,))
        connection.commit()
        return None

    return user


def get_optional_current_user(request: Request) -> dict[str, Any] | None:
    session_token = request.cookies.get(SESSION_COOKIE_NAME)
    with get_connection() as connection:
        return get_session_user(connection, session_token)


def require_authenticated_user(request: Request) -> dict[str, Any]:
    user = get_optional_current_user(request)
    if user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="AUTH_REQUIRED")
    return user


def require_app_user(request: Request) -> dict[str, Any]:
    user = require_authenticated_user(request)
    if parse_bool(user.get("must_change_password")):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PASSWORD_CHANGE_REQUIRED")
    return user


def require_owner_user(request: Request) -> dict[str, Any]:
    user = require_app_user(request)
    if str(user.get("role")) != ROLE_OWNER:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="OWNER_REQUIRED")
    return user


def set_session_cookie(response: Response, session_token: str) -> None:
    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=session_token,
        httponly=True,
        samesite="lax",
        secure=False,
        path="/",
    )


def clear_session_cookie(response: Response) -> None:
    response.delete_cookie(key=SESSION_COOKIE_NAME, path="/")


def normalize_country_code(country: str | None) -> str:
    normalized = (country or DEFAULT_COUNTRY).strip().lower()
    if normalized not in SUPPORTED_COUNTRIES:
        raise HTTPException(status_code=400, detail="Неизвестная страна.")
    return normalized


def load_custom_courts(connection, country: str) -> list[dict[str, str]]:
    rows = connection.execute(
        """
        SELECT name, city, region
        FROM custom_courts
        WHERE COALESCE(country, ?) = ?
        ORDER BY name COLLATE NOCASE ASC
        """,
        (DEFAULT_COUNTRY, normalize_country_code(country)),
    ).fetchall()
    result: list[dict[str, str]] = []
    for row in rows:
        item = dict(row)
        values = [str(item.get("name") or "").strip(), str(item.get("city") or "").strip(), str(item.get("region") or "").strip()]
        if not all(values):
            continue
        if any("?" in value for value in values):
            continue
        result.append(item)
    return result


def build_reference_catalog(connection, country: str) -> dict[str, Any]:
    normalized_country = normalize_country_code(country)
    custom_courts = load_custom_courts(connection, normalized_country)
    if normalized_country == "uz":
        return merge_static_court_catalog(UZBEKISTAN_COURT_CATALOG, custom_courts)
    return merge_court_catalog(custom_courts)


def merge_static_court_catalog(
    base_catalog: dict[str, Any], custom_courts: list[dict[str, str]]
) -> dict[str, Any]:
    regions = list(base_catalog["regions"])
    cities = set(base_catalog["cities"])
    courts_by_city = {city: list(items) for city, items in base_catalog["courtsByCity"].items()}
    courts_by_region = {region: list(items) for region, items in base_catalog["courtsByRegion"].items()}
    court_city_map = dict(base_catalog["courtCityMap"])
    court_region_map = dict(base_catalog["courtRegionMap"])
    city_region_map = dict(base_catalog["cityRegionMap"])
    cities_by_region = {region: list(items) for region, items in base_catalog["citiesByRegion"].items()}

    for item in custom_courts:
        name = str(item["name"]).strip()
        city = str(item["city"]).strip()
        region = str(item["region"]).strip()
        if not name or not city or not region:
            continue

        if region not in regions:
            regions.append(region)
        cities.add(city)
        court_city_map[name] = city
        court_region_map[name] = region
        city_region_map[city] = region
        courts_by_city.setdefault(city, [])
        if name not in courts_by_city[city]:
            courts_by_city[city].append(name)
        courts_by_region.setdefault(region, [])
        if name not in courts_by_region[region]:
            courts_by_region[region].append(name)
        cities_by_region.setdefault(region, [])
        if city not in cities_by_region[region]:
            cities_by_region[region].append(city)

    return {
        "regions": sorted(regions),
        "cities": sorted(cities),
        "courtsByCity": {city: sorted(set(items)) for city, items in courts_by_city.items()},
        "courtsByRegion": {region: sorted(set(items)) for region, items in courts_by_region.items()},
        "courtCityMap": court_city_map,
        "courtRegionMap": court_region_map,
        "cityRegionMap": city_region_map,
        "citiesByRegion": {region: sorted(set(items)) for region, items in cities_by_region.items()},
    }


def align_city_with_court(connection, country: str, city: str, court: str) -> tuple[str, str]:
    normalized_city = city.strip()
    normalized_court = court.strip()
    catalog = build_reference_catalog(connection, country)
    mapped_city = catalog["courtCityMap"].get(normalized_court)
    if mapped_city:
        return mapped_city, normalized_court
    return normalized_city, normalized_court


def find_requisites_path() -> Path:
    for candidate in REQUISITES_PATH_CANDIDATES:
        if candidate.exists():
            return candidate

    for candidate in PROJECTS_DIR.glob("*еквизит*.xlsx"):
        if candidate.exists():
            return candidate

    raise FileNotFoundError("Файл с реквизитами компаний не найден.")


def normalize_company_key(value: str | None) -> str:
    text = str(value or "").strip()
    translation = str.maketrans(
        {
            "А": "A",
            "В": "B",
            "Е": "E",
            "К": "K",
            "М": "M",
            "Н": "H",
            "О": "O",
            "Р": "P",
            "С": "C",
            "Т": "T",
            "Х": "X",
            "Ё": "E",
            "І": "I",
            "Ї": "I",
            "Ъ": "",
            "Ь": "",
            "“": '"',
            "”": '"',
            "«": '"',
            "»": '"',
        }
    )
    normalized = text.upper().translate(translation)
    return "".join(ch for ch in normalized if ch.isalnum())


STATIC_COMPANY_REQUISITES: dict[str, dict[str, str]] = {
    normalize_company_key("ООО «Concept Evolution»"): {
        "company_name": "ООО «Concept Evolution»",
        "company_block": "ООО «Concept Evolution»\nг. Ташкент, Шайхонтахурский район, ул. Лабзак, 64-а",
        "director_name": "Умарова Хусния Бону Абдугофур кизи",
        "bank_name": 'JSC "KAPITALBANK"',
        "iik": "",
        "bik": "",
        "bin": "312040649",
        "address": "г. Ташкент, Шайхонтахурский район, ул. Лабзак, 64-а",
        "kbe": "",
        "account_number": "2020 8000 9072 3076 2001",
        "bank_mfo": "01158",
    },
    normalize_company_key("ООО «RRS RETAIL CITY»"): {
        "company_name": "ООО «RRS RETAIL CITY»",
        "company_block": "ООО «RRS RETAIL CITY»\nгород Ташкент, Мирзо Улугбекский район, МСГ Элобод, улица Чуст, дом 1.",
        "director_name": "Кульмедова Ася Бекчановна",
        "bank_name": 'JSC "KAPITALBANK"',
        "iik": "",
        "bik": "",
        "bin": "311463803",
        "address": "город Ташкент, Мирзо Улугбекский район, МСГ Элобод, улица Чуст, дом 1.",
        "kbe": "",
        "account_number": "2020 8000 4071 0632 2001",
        "bank_mfo": "01158",
    },
}


@lru_cache(maxsize=1)
def load_company_requisites() -> dict[str, dict[str, str]]:
    path = find_requisites_path()
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(value or "").strip() for value in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers)}

    def get_value(row: tuple[Any, ...], column_name: str) -> str:
        idx = header_index.get(column_name)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx] or "").strip()

    records: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        company_name = get_value(row, headers[0])
        if not company_name:
            continue

        record = {
            "company_name": company_name,
            "company_block": get_value(row, "Реквизиты для правого угла верхнего"),
            "director_name": get_value(row, "ФИО директора"),
            "bank_name": get_value(row, "Наименование банка"),
            "iik": get_value(row, "ИИК"),
            "bik": get_value(row, "БИК"),
            "bin": get_value(row, "БИН"),
            "address": get_value(row, "Адрес"),
            "kbe": "",
            "account_number": get_value(row, "Р /СЃ") or get_value(row, "Р Р°СЃС‡РµС‚РЅС‹Р№ СЃС‡РµС‚"),
            "bank_mfo": get_value(row, "РњР¤Рћ") or get_value(row, "MFO"),
        }
        records[normalize_company_key(company_name)] = record

    records.update(STATIC_COMPANY_REQUISITES)

    return records


def find_company_requisites(company_name: str) -> dict[str, str] | None:
    company_key = normalize_company_key(company_name)
    requisites = load_company_requisites()
    exact = requisites.get(company_key)
    if exact is not None:
        return exact

    normalized_lookup_key = normalize_company_lookup_key(company_name)
    for key, record in requisites.items():
        if normalize_company_lookup_key(key) == normalized_lookup_key:
            return record
    return None


def normalize_company_lookup_key(value: str | None) -> str:
    key = normalize_company_key(value)
    for prefix in ("TOO", "T00"):
        if key.startswith(prefix):
            key = key[len(prefix) :]
    return key


def match_company_to_library(company_name: str, country: str = DEFAULT_COUNTRY) -> str:
    sanitized_source = sanitize_company_name(company_name)
    normalized_source = normalize_company_lookup_key(sanitized_source)
    if not normalized_source:
        return sanitized_source

    exact_match: str | None = None
    partial_match: str | None = None

    for candidate in get_companies_by_country(normalize_country_code(country)):
        normalized_candidate = normalize_company_lookup_key(candidate)
        if normalized_candidate == normalized_source:
            exact_match = candidate
            break
        if (
            normalized_source in normalized_candidate
            or normalized_candidate in normalized_source
        ) and partial_match is None:
            partial_match = candidate

    return exact_match or partial_match or sanitized_source


def sanitize_company_name(company_name: str | None) -> str:
    text = normalize_text(company_name) or ""
    lowered = text.lower()
    if "boggner" in lowered or "bogner" in lowered or "??? ?boggner?" in lowered:
        return "\u0422\u041e\u041e \u00abBoggner\u00bb"
    return text


def _match_company_to_library_legacy(company_name: str, country: str = DEFAULT_COUNTRY) -> str:
    normalized_source = normalize_company_lookup_key(company_name)
    if not normalized_source:
        return company_name

    exact_match: str | None = None
    partial_match: str | None = None

    for candidate in get_companies_by_country(normalize_country_code(country)):
        normalized_candidate = normalize_company_lookup_key(candidate)
        if normalized_candidate == normalized_source:
            exact_match = candidate
            break
        if (
            normalized_source in normalized_candidate
            or normalized_candidate in normalized_source
        ) and partial_match is None:
            partial_match = candidate

    return exact_match or partial_match or company_name


def build_short_client_name(full_name: str) -> str:
    parts = [part for part in (normalize_text(full_name) or "").split(" ") if part]
    if not parts:
        return "—"
    surname = parts[0]
    initials = " ".join(f"{part[0]}." for part in parts[1:] if part)
    return " ".join(part for part in [surname, initials] if part)


def add_months_preserving_day(value: date, months: int) -> date:
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    day = min(value.day, monthrange(year, month)[1])
    return date(year, month, day)


def build_lawsuit_penalty_rows(
    start_date: date,
    end_date: date,
    claim_sent_date: date,
    monthly_payment_amount: float,
    first_period_paid_amount: float = 0,
) -> tuple[list[dict[str, Any]], float, int]:
    if end_date <= start_date:
        raise HTTPException(status_code=400, detail="Период рассрочки заполнен некорректно.")
    if claim_sent_date <= start_date:
        raise HTTPException(status_code=400, detail="Дата претензии должна быть позже начала рассрочки.")

    rows: list[dict[str, Any]] = []
    current_start = start_date
    period_index = 1
    first_outstanding = max(round(monthly_payment_amount - first_period_paid_amount, 2), 0.0)
    schedule_end = min(end_date, claim_sent_date)

    while current_start < schedule_end:
        current_end = min(add_months_preserving_day(current_start, 1), schedule_end)
        days = max((current_end - current_start).days, 0)
        obligation = round(first_outstanding + monthly_payment_amount * max(period_index - 1, 0), 2)
        penalty_amount = round(obligation * 0.001 * days, 2)
        rows.append(
            {
                "period_from": current_start,
                "period_to": current_end,
                "days": days,
                "obligation_amount": obligation,
                "penalty_rate_percent": 0.1,
                "penalty_amount": penalty_amount,
            }
        )
        current_start = current_end
        period_index += 1

    if claim_sent_date > end_date:
        days = max((claim_sent_date - end_date).days, 0)
        # By the time the installment schedule ends, the next unpaid monthly tranche
        # has already matured and must be included in the final post-schedule period.
        final_obligation = round(first_outstanding + monthly_payment_amount * max(period_index - 1, 0), 2)
        penalty_amount = round(final_obligation * 0.001 * days, 2)
        rows.append(
            {
                "period_from": end_date,
                "period_to": claim_sent_date,
                "days": days,
                "obligation_amount": final_obligation,
                "penalty_rate_percent": 0.1,
                "penalty_amount": penalty_amount,
            }
        )

    computed_total = round(sum(row["penalty_amount"] for row in rows), 2)
    total_days = sum(row["days"] for row in rows)
    return rows, computed_total, total_days


def find_pdf_browser_path() -> Path:
    for candidate in PDF_BROWSER_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("Не найден локальный браузер для генерации PDF.")


def preferred_phone(row: dict[str, Any]) -> str:
    return normalize_text(row.get("mobile_phone")) or normalize_text(row.get("home_phone")) or "—"


def format_money(value: float | int | None) -> str:
    amount = round(float(value or 0), 2)
    integer_part = int(amount)
    decimal_part = int(round((amount - integer_part) * 100))
    integer_display = f"{integer_part:,}".replace(",", " ")
    return f"{integer_display},{decimal_part:02d}"


UNITS_MALE = (
    "",
    "один",
    "два",
    "три",
    "четыре",
    "пять",
    "шесть",
    "семь",
    "восемь",
    "девять",
)
UNITS_FEMALE = (
    "",
    "одна",
    "две",
    "три",
    "четыре",
    "пять",
    "шесть",
    "семь",
    "восемь",
    "девять",
)
TEENS = (
    "десять",
    "одиннадцать",
    "двенадцать",
    "тринадцать",
    "четырнадцать",
    "пятнадцать",
    "шестнадцать",
    "семнадцать",
    "восемнадцать",
    "девятнадцать",
)
TENS = (
    "",
    "",
    "двадцать",
    "тридцать",
    "сорок",
    "пятьдесят",
    "шестьдесят",
    "семьдесят",
    "восемьдесят",
    "девяносто",
)
HUNDREDS = (
    "",
    "сто",
    "двести",
    "триста",
    "четыреста",
    "пятьсот",
    "шестьсот",
    "семьсот",
    "восемьсот",
    "девятьсот",
)
GROUP_FORMS = (
    ("", "", "", False),
    ("тысяча", "тысячи", "тысяч", True),
    ("миллион", "миллиона", "миллионов", False),
    ("миллиард", "миллиарда", "миллиардов", False),
)


def choose_plural(value: int, forms: tuple[str, str, str]) -> str:
    value = abs(value) % 100
    if 11 <= value <= 19:
        return forms[2]
    last_digit = value % 10
    if last_digit == 1:
        return forms[0]
    if 2 <= last_digit <= 4:
        return forms[1]
    return forms[2]


def triplet_to_words(value: int, *, female: bool) -> list[str]:
    words: list[str] = []
    words.append(HUNDREDS[value // 100])
    remainder = value % 100
    if 10 <= remainder <= 19:
        words.append(TEENS[remainder - 10])
    else:
        words.append(TENS[remainder // 10])
        units = UNITS_FEMALE if female else UNITS_MALE
        words.append(units[remainder % 10])
    return [word for word in words if word]


def number_to_words_ru(value: int) -> str:
    if value == 0:
        return "ноль"

    words: list[str] = []
    group_index = 0
    while value > 0:
        value, chunk = divmod(value, 1000)
        if chunk:
            forms = GROUP_FORMS[group_index] if group_index < len(GROUP_FORMS) else GROUP_FORMS[-1]
            words = triplet_to_words(chunk, female=forms[3]) + (
                [choose_plural(chunk, forms[:3])] if forms[0] else []
            ) + words
        group_index += 1
    return " ".join(words)


def money_to_words_ru(value: float | int | None) -> str:
    amount = round(float(value or 0), 2)
    integer_part = int(amount)
    decimal_part = int(round((amount - integer_part) * 100))
    tenge_words = choose_plural(integer_part, ("тенге", "тенге", "тенге"))
    if decimal_part:
        tiyn_words = choose_plural(decimal_part, ("тиын", "тиына", "тиын"))
        return f"{number_to_words_ru(integer_part)} {tenge_words} {decimal_part:02d} {tiyn_words}"
    return f"{number_to_words_ru(integer_part)} {tenge_words}"


def money_to_words_sum_ru(value: float | int | None) -> str:
    amount = round(float(value or 0), 2)
    integer_part = int(amount)
    decimal_part = int(round((amount - integer_part) * 100))
    sum_words = choose_plural(integer_part, ("сум", "сума", "сум"))
    if decimal_part:
        return f"{number_to_words_ru(integer_part)} {sum_words} {decimal_part:02d} тийин"
    return f"{number_to_words_ru(integer_part)} {sum_words}"


def compute_lawsuit_claim_price(debt_amount: float, penalty_amount: float) -> float:
    return round(float(debt_amount or 0) + float(penalty_amount or 0), 2)


def compute_lawsuit_state_duty(country: str, debt_amount: float, penalty_amount: float) -> float:
    claim_price_amount = compute_lawsuit_claim_price(debt_amount, penalty_amount)
    if normalize_country_code(country) == "uz":
        return max(round(claim_price_amount * 0.04, 2), UZ_BRV_AMOUNT)
    return round(claim_price_amount * 0.03, 2)


def compute_simple_penalty_amount(debt_amount: float, overdue_days: int) -> float:
    return round(max(float(debt_amount or 0), 0.0) * 0.001 * max(int(overdue_days or 0), 0), 2)


def build_company_payment_details(requisites: dict[str, str]) -> str:
    parts = [
        requisites.get("bin"),
        requisites.get("bank_name"),
        f"ИИК: {requisites.get('iik')}" if requisites.get("iik") else "",
        f"БИК: {requisites.get('bik')}" if requisites.get("bik") else "",
        f"КБЕ: {requisites.get('kbe') or '—'}",
    ]
    return ", ".join(part for part in parts if part)


def build_company_payment_detail_lines(requisites: dict[str, str]) -> list[str]:
    lines: list[str] = []
    if requisites.get("bin"):
        lines.append(str(requisites["bin"]))
    if requisites.get("bank_name"):
        lines.append(f"Наименование банка: {requisites['bank_name']}")
    if requisites.get("iik"):
        lines.append(f"ИИК: {requisites['iik']}")
    if requisites.get("bik"):
        lines.append(f"БИК: {requisites['bik']}")
    return lines


def build_company_payment_detail_lines_uz(requisites: dict[str, str]) -> list[str]:
    lines: list[str] = []
    if requisites.get("bin"):
        lines.append(f"ИНН: {requisites['bin']}")
    if requisites.get("bank_name"):
        lines.append(f"Наименование банка: {requisites['bank_name']}")
    if requisites.get("account_number"):
        lines.append(f"Расчетный счет: {requisites['account_number']}")
    if requisites.get("bank_mfo"):
        lines.append(f"МФО банка: {requisites['bank_mfo']}")
    return lines


def normalize_document_products(
    raw_products: list[dict[str, Any]] | None,
    *,
    fallback_name: str,
) -> list[dict[str, Any]]:
    products: list[dict[str, Any]] = []
    for item in raw_products or []:
        name = normalize_text(item.get("name")) or normalize_text(item.get("display_name"))
        if not name:
            continue
        quantity_raw = item.get("quantity", 1)
        try:
            quantity = int(float(quantity_raw))
        except (TypeError, ValueError):
            quantity = 1
        quantity = max(quantity, 1)
        products.append(
            {
                "name": name,
                "quantity": quantity,
                "display_name": f"{name} — {quantity} шт.",
            }
        )
    if not products:
        products.append(
            {
                "name": fallback_name,
                "quantity": 1,
                "display_name": f"{fallback_name} — 1 шт.",
            }
        )
    return products


def build_company_header_lines(requisites: dict[str, str], company_name: str) -> list[str]:
    company_block = normalize_text(requisites.get("company_block"))
    if company_block:
        lines = [line.strip() for line in company_block.splitlines() if line.strip()]
        if lines:
            return lines

    address = normalize_text(requisites.get("address"))
    return [line for line in [company_name, address] if line] or [company_name or "—"]


def load_font(size: int, *, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_path = SERIF_FONT_BOLD if bold else SERIF_FONT_REGULAR
    try:
        return ImageFont.truetype(str(font_path), size=size)
    except OSError:
        return ImageFont.load_default()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    normalized = " ".join((text or "").split())
    if not normalized:
        return [""]

    words = normalized.split(" ")
    lines: list[str] = []
    current = words[0]

    for word in words[1:]:
        candidate = f"{current} {word}".strip()
        if draw.textlength(candidate, font=font) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word

    lines.append(current)
    return lines


def draw_text_block(
    draw: ImageDraw.ImageDraw,
    text: str,
    *,
    x: int,
    y: int,
    width: int,
    font: ImageFont.ImageFont,
    line_spacing: int,
    align: str = "left",
    paragraph_spacing: int = 0,
    first_line_indent: int = 0,
) -> int:
    normalized = " ".join((text or "").split())
    if not normalized:
        return y + paragraph_spacing

    if first_line_indent > 0:
        words = normalized.split(" ")
        first_line_words: list[str] = []
        while words:
            candidate_words = first_line_words + [words[0]]
            candidate = " ".join(candidate_words).strip()
            if draw.textlength(candidate, font=font) <= max(40, width - first_line_indent):
                first_line_words.append(words.pop(0))
            else:
                break

        first_line = " ".join(first_line_words).strip()
        remaining = " ".join(words).strip()
        lines = [first_line] if first_line else []
        if remaining:
            lines.extend(wrap_text(draw, remaining, font, width))
    else:
        lines = wrap_text(draw, normalized, font, width)

    bbox = draw.textbbox((0, 0), "Аг", font=font)
    line_height = (bbox[3] - bbox[1]) + line_spacing

    for index, line in enumerate(lines):
        extra_indent = first_line_indent if index == 0 and align == "left" else 0
        line_width = draw.textlength(line, font=font)
        if align == "center":
            line_x = x + max(0, (width - line_width) / 2)
        elif align == "right":
            line_x = x + max(0, width - line_width)
        else:
            line_x = x + extra_indent

        draw.text((line_x, y), line, fill="#111111", font=font)
        y += line_height

    return y + paragraph_spacing


def _render_claim_pdf_kz(
    debtor: dict[str, Any],
    *,
    debt_amount_override: float | None = None,
    product_overrides: list[dict[str, Any]] | None = None,
) -> Path:
    company_name_value = sanitize_company_name(str(debtor.get("company") or ""))
    requisites = find_company_requisites(company_name_value)
    if requisites is None:
        raise HTTPException(
            status_code=400,
            detail=f"Для компании «{company_name_value or '—'}» не найдены реквизиты.",
        )

    contract_number = str(debtor.get("contract_number") or "")
    contract_date = contract_date_value_from_number(contract_number)
    financials = compute_financials(debtor)
    contract_total_amount = debtor.get("contract_total_amount")
    contract_advance_amount = debtor.get("contract_advance_amount")
    if contract_total_amount in (None, ""):
        raise HTTPException(
            status_code=400,
            detail="Не заполнена сумма контракта. Сначала подтяните или заполните значение договора.",
        )
    company_name = requisites.get("company_name") or company_name_value or "—"
    contract_total_amount = parse_float(contract_total_amount)
    debt_amount = float(debt_amount_override) if debt_amount_override is not None else financials["debt_amount"]
    if contract_advance_amount in (None, ""):
        contract_advance_amount = max(round(contract_total_amount - financials["debt_amount"], 2), 0.0)
    else:
        contract_advance_amount = parse_float(contract_advance_amount)

    try:
        crm_context = DiSellApiClient().lookup_lawsuit_context(
            contract_number,
            country=str(debtor.get("country") or DEFAULT_COUNTRY),
        )
    except DiSellApiError:
        crm_context = {}

    products = normalize_document_products(
        product_overrides or crm_context.get("products"),
        fallback_name="Товар по договору",
    )
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    safe_contract = re.sub(r"[^A-Za-z0-9_-]+", "_", str(debtor.get("contract_number") or debtor.get("id")))
    pdf_path = GENERATED_DIR / f"claim_{safe_contract}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    image_width = 1654
    image_height = 2339
    margin_left = 128
    margin_right = 128
    margin_top = 82
    margin_bottom = 82
    content_width = image_width - margin_left - margin_right
    right_block_width = 520
    title_width = content_width
    content_x = margin_left
    paragraph_indent = 44

    regular_font = load_font(28)
    title_font = load_font(34, bold=True)
    ask_font = load_font(31, bold=True)
    small_font = load_font(24)

    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)

    y = margin_top
    right_x = content_x + content_width - right_block_width
    right_lines = [
        debtor.get("client_name") or "—",
        normalize_text(debtor.get("address")) or "—",
        preferred_phone(debtor),
    ]
    for line in right_lines:
        y = draw_text_block(
            draw,
            str(line),
            x=right_x,
            y=y,
            width=right_block_width,
            font=regular_font,
            line_spacing=12,
            align="left",
            paragraph_spacing=2,
        )

    y += 12
    for line in build_company_header_lines(requisites, company_name):
        y = draw_text_block(
            draw,
            line,
            x=right_x,
            y=y,
            width=right_block_width,
            font=regular_font,
            line_spacing=12,
            align="left",
            paragraph_spacing=1,
        )

    y += 20
    y = draw_text_block(
        draw,
        "Досудебная претензия",
        x=content_x,
        y=y,
        width=title_width,
        font=title_font,
        line_spacing=12,
        align="center",
        paragraph_spacing=28,
    )

    paragraphs = [
        (
            f"Между {company_name} и Вами заключен договор купли-продажи товара № "
            f"{contract_number or '—'} от {format_date(contract_date) or '—'}, "
            f"в соответствии с условиями которого {company_name} передала в собственность "
            "покупателя товар:"
        ),
        (
            f"Согласно условиям заключенного договора покупатель принял на себя обязательства "
            f"по оплате стоимости полученного товара. Стоимость переданного товара составила "
            f"{format_money(contract_total_amount)} тенге, из которых "
            f"{format_money(contract_advance_amount)} тенге были оплачены в качестве "
            "первоначального взноса. Оставшаяся сумма подлежала оплате в соответствии с "
            f"графиком платежей, предусмотренным пунктом 2.1 договора купли-продажи № "
            f"{contract_number or '—'}."
        ),
        (
            "В соответствии со статьей 406 Гражданского кодекса Республики Казахстан по "
            "договору купли-продажи одна сторона (продавец) обязуется передать имущество "
            "(товар) в собственность, хозяйственное ведение либо оперативное управление "
            "другой стороне (покупателю), а покупатель обязуется принять указанный товар "
            "и уплатить за него определенную денежную сумму (цену)."
        ),
        (
            "Согласно статье 272 Гражданского кодекса Республики Казахстан обязательства "
            "должны исполняться надлежащим образом в соответствии с условиями обязательства "
            "и требованиями законодательства, а при отсутствии таких условий и требований — "
            "в соответствии с обычаями делового оборота и иными обычно предъявляемыми "
            "требованиями."
        ),
        (
            "В силу статьи 277 Гражданского кодекса Республики Казахстан, если обязательство "
            "предусматривает или позволяет определить день его исполнения либо период времени, "
            "в течение которого оно должно быть исполнено, обязательство подлежит исполнению "
            "в установленный срок."
        ),
        (
            f"Однако принятые на себя обязательства Вами исполнены не были, что привело к "
            f"нарушению прав и законных интересов {company_name}."
        ),
        (
            "Кроме того, условиями заключенного между сторонами договора предусмотрено, что "
            "в случае просрочки очередного платежа при оплате товара в рассрочку покупатель "
            "обязан уплатить продавцу неустойку (пеню) в размере 0,1% за каждый день "
            "просрочки от суммы платежа, подлежащего оплате."
        ),
    ]

    for paragraph in paragraphs:
        y = draw_text_block(
            draw,
            paragraph,
            x=content_x,
            y=y,
            width=content_width,
            font=regular_font,
            line_spacing=14,
            align="left",
            paragraph_spacing=12,
            first_line_indent=paragraph_indent,
        )

    for product in products:
        display_name = normalize_text(product.get("display_name")) or normalize_text(product.get("name")) or "Товар"
        y = draw_text_block(
            draw,
            display_name,
            x=content_x,
            y=y,
            width=content_width,
            font=regular_font,
            line_spacing=14,
            align="left",
            paragraph_spacing=6,
            first_line_indent=0,
        )

    y += 24
    y = draw_text_block(
        draw,
        "ПРОШУ:",
        x=content_x,
        y=y,
        width=content_width,
        font=ask_font,
        line_spacing=12,
        align="center",
        paragraph_spacing=24,
    )

    payment_detail_lines = build_company_payment_detail_lines(requisites)

    ask_paragraphs = [
        (
            f"В течение 7 (семи) календарных дней с момента получения настоящей претензии "
            f"осуществить в пользу {company_name} выплату суммы задолженности в размере "
            f"{format_money(debt_amount)} ({money_to_words_ru(debt_amount)}) тенге по следующим реквизитам:"
        ),
        "Настоящая досудебная претензия направляется в рамках претензионно-исковой работы.",
        (
            f"В случае неудовлетворения требований {company_name}, выражающихся в погашении "
            "задолженности в добровольном порядке, "
            f"{company_name} будет вынуждена обратиться в судебные органы за защитой своих "
            "прав и законных интересов с возложением на Вас дополнительных расходов, включая "
            "неустойку, государственную пошлину, а также расходы на оплату услуг представителя."
        ),
        "Надеемся на понимание и урегулирование сложившейся ситуации в добровольном порядке.",
    ]

    for index, paragraph in enumerate(ask_paragraphs):
        y = draw_text_block(
            draw,
            paragraph,
            x=content_x,
            y=y,
            width=content_width,
            font=regular_font,
            line_spacing=14,
            align="left",
            paragraph_spacing=12 if index != 1 else 18,
            first_line_indent=paragraph_indent if index != 1 else 0,
        )

        if index == 0 and payment_detail_lines:
            for payment_line in payment_detail_lines:
                y = draw_text_block(
                    draw,
                    payment_line,
                    x=content_x,
                    y=y,
                    width=content_width,
                    font=regular_font,
                    line_spacing=14,
                    align="left",
                    paragraph_spacing=4,
                    first_line_indent=0,
                )
            y += 10

    signature_y = max(y + 16, image_height - margin_bottom - 90)
    director_name = requisites.get("director_name") or "—"
    signature_text = "Директор"
    draw.text((margin_left, signature_y), signature_text, fill="#111111", font=small_font)

    label_width = draw.textlength(signature_text, font=small_font)
    line_start_x = int(margin_left + label_width + 18)
    line_end_x = line_start_x + 200
    line_y = signature_y + 24
    draw.line((line_start_x, line_y, line_end_x, line_y), fill="#111111", width=1)

    draw.text((line_end_x + 18, signature_y), director_name, fill="#111111", font=small_font)

    generated_date = format_date(date.today()) or date.today().isoformat()
    date_bbox = draw.textbbox((0, 0), generated_date, font=small_font)
    date_width = date_bbox[2] - date_bbox[0]
    draw.text((image_width - margin_right - date_width, signature_y), generated_date, fill="#111111", font=small_font)

    try:
        image.save(pdf_path, "PDF", resolution=150.0)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF претензии.") from exc

    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF претензии.")

    return pdf_path


def _render_claim_pdf_uz(
    debtor: dict[str, Any],
    *,
    debt_amount_override: float | None = None,
    product_overrides: list[dict[str, Any]] | None = None,
) -> Path:
    company_name_value = match_company_to_library(
        sanitize_company_name(str(debtor.get("company") or "")),
        "uz",
    )
    requisites = find_company_requisites(company_name_value)
    if requisites is None:
        raise HTTPException(
            status_code=400,
            detail=f"Р”Р»СЏ РєРѕРјРїР°РЅРёРё В«{company_name_value or '—'}В» РЅРµ РЅР°Р№РґРµРЅС‹ СЂРµРєРІРёР·РёС‚С‹.",
        )

    try:
        crm_context = DiSellApiClient().lookup_lawsuit_context(
            str(debtor.get("contract_number") or ""),
            country="uz",
        )
    except DiSellApiError:
        crm_context = {}
    if debt_amount_override is None and crm_context.get("debt_amount") not in (None, ""):
        debt_amount = parse_float(crm_context.get("debt_amount"))

    company_name = requisites.get("company_name") or company_name_value or "—"
    contract_number = str(crm_context.get("contract_number") or debtor.get("contract_number") or "—")
    contract_date = contract_date_value_from_number(contract_number)
    crm_contract_date = crm_context.get("contract_date")
    if isinstance(crm_contract_date, str) and crm_contract_date:
        try:
            contract_date = date.fromisoformat(crm_contract_date)
        except ValueError:
            pass

    financials = compute_financials(debtor)
    contract_total_amount = crm_context.get("contract_total_amount")
    if contract_total_amount in (None, ""):
        contract_total_amount = debtor.get("contract_total_amount")
    if contract_total_amount in (None, ""):
        raise HTTPException(
            status_code=400,
            detail="РќРµ Р·Р°РїРѕР»РЅРµРЅР° СЃСѓРјРјР° РєРѕРЅС‚СЂР°РєС‚Р°.",
        )
    contract_total_amount = parse_float(contract_total_amount)

    contract_advance_amount = crm_context.get("advance_amount")
    if contract_advance_amount in (None, ""):
        contract_advance_amount = debtor.get("contract_advance_amount")
    if contract_advance_amount in (None, ""):
        contract_advance_amount = max(round(contract_total_amount - financials["debt_amount"], 2), 0.0)
    else:
        contract_advance_amount = parse_float(contract_advance_amount)

    debt_amount = float(debt_amount_override) if debt_amount_override is not None else financials["debt_amount"]
    products = normalize_document_products(
        product_overrides or crm_context.get("products"),
        fallback_name="Товар по договору",
    )

    client_name = normalize_text(crm_context.get("client_name")) or normalize_text(debtor.get("client_name")) or "—"
    client_address = normalize_text(crm_context.get("client_address")) or normalize_text(debtor.get("address")) or "—"
    crm_phones = crm_context.get("client_phones") or []
    normalized_phones = [
        phone
        for phone in (
            [normalize_text(phone) for phone in crm_phones]
            + [normalize_text(debtor.get("mobile_phone")), normalize_text(debtor.get("home_phone"))]
        )
        if phone
    ]
    client_phone = ", ".join(dict.fromkeys(normalized_phones)) if normalized_phones else "—"

    debt_words_ru = money_to_words_sum_ru(debt_amount)
    payment_detail_lines_ru = build_company_payment_detail_lines_uz(requisites)
    payment_detail_lines_uz = []
    if requisites.get("bin"):
        payment_detail_lines_uz.append(f"ИНН: {requisites['bin']}")
    if requisites.get("bank_name"):
        payment_detail_lines_uz.append(f"Банк номи: {requisites['bank_name']}")
    if requisites.get("account_number"):
        payment_detail_lines_uz.append(f"Ҳисоб рақами: {requisites['account_number']}")
    if requisites.get("bank_mfo"):
        payment_detail_lines_uz.append(f"Банк МФОси: {requisites['bank_mfo']}")

    ru_paragraphs = [
        f"Между {company_name} и Вами заключен договор купли/продажи товара № {contract_number} от {format_date(contract_date) or '—'}, в соответствии с которым {company_name} в собственность покупателя передан товар:",
        f"Согласно условиям заключенного договора определено, что покупатель принял на себя обязательства по оплате стоимости полученного товара. Стоимость переданного товара была определена в размере {format_money(contract_total_amount)} сум, из которых {format_money(contract_advance_amount)} сум были оплачены в качестве предоплаты, остальная оплата по договору осуществлялась согласно графика указанного в 2.2 договора купли-продажи.",
        "В соответствии со статьёй 386 Гражданского кодекса Республики Узбекистан, по договору купли-продажи одна сторона (продавец) обязуется передать товар в собственность другой стороне (покупателю), а покупатель — принять товар и уплатить за него установленную договором денежную сумму.",
        "В силу статьи 236 Гражданского кодекса Республики Узбекистан, обязательства должны исполняться надлежащим образом в соответствии с условиями обязательства и требованиями законодательства.",
        f"Однако обязательства, принятые на себя, Вы не исполнили, что привело к нарушению прав и охраняемых законом интересов {company_name}.",
        "Одновременно с этим, пунктом 2.3 заключенного договора определено, что в случае просрочки очередного платежа при оплате товара в рассрочку покупатель уплачивает продавцу пеню в размере 0,1 % за каждый день просрочки от суммы платежа, подлежащего уплате.",
    ]

    uz_paragraphs = [
        f"{company_name} ва Сиз ўртасида {format_date(contract_date) or '—'} санадаги {contract_number}-сонли товар олди-сотди шартномаси тузилган бўлиб, унга кўра {company_name} харидорга қуйидаги товарларни топширган:",
        f"Шартнома шартларига кўра, харидор олинган товар қийматини тўлаш мажбуриятини олган. Товар қиймати {format_money(contract_total_amount)} сўм этиб белгиланган, шундан {format_money(contract_advance_amount)} сўм олдиндан тўлов сифатида тўланган, қолган қисми эса олди-сотди шартномасининг 2.2-бандида кўрсатилган жадвал асосида тўланиши лозим бўлган.",
        "Ўзбекистон Республикаси Фуқаролик кодексининг 386-моддасига кўра, олди-сотди шартномаси бўйича сотувчи товарни харидорга мулк қилиб топшириши, харидор эса товарни қабул қилиб, белгиланган пул суммасини тўлаши шарт.",
        "Фуқаролик кодексининг 236-моддасига мувофиқ, мажбуриятлар шартнома шартлари ва қонунчилик талабларига мувофиқ лозим даражада бажарилиши керак.",
        f"Бироқ Сиз зиммангизга олган мажбуриятларни бажармагансиз, бу эса {company_name}нинг ҳуқуқлари ва қонуний манфаатлари бузилишига олиб келган.",
        "Шунингдек, шартноманинг 2.3-бандига кўра, муддатли тўлов бўйича навбатдаги тўлов кечиктирилган тақдирда, харидор тўланиши лозим бўлган суммадан ҳар бир кечиктирилган кун учун 0,1 % миқдорида пеня тўлайди.",
    ]

    ru_ask = [
        f"В течении 7 дней осуществить в пользу {company_name} выплату суммы задолженности в размере {format_money(debt_amount)} ({debt_words_ru}) сум по следующим реквизитам:",
        "Сообщаем, что указанная досудебная претензия направлена в рамках претензионно-исковой работы. В случае неудовлетворения требования, указанного в настоящей претензии, компания будет вынуждена обратиться в суд за защитой своих интересов, с возложением на Вас дополнительных расходов по уплате неустойки, государственной пошлины и услуг представителя.",
        "Мы надеемся на добросовестный подход к исполнению обязательств и предлагаем урегулировать вопрос в досудебном порядке.",
    ]
    uz_ask = [
        f"7 кун ичида {company_name} фойдасига {format_money(debt_amount)} сўм миқдоридаги қарздорликни қуйидаги реквизитлар бўйича тўлашингизни сўраймиз:",
        "Мазкур судгача бўлган талабнома даъво ишларини юритиш доирасида юборилмоқда. Агар ундаги талаблар бажарилмаса, компания ўз манфаатларини ҳимоя қилиш учун судга мурожаат қилишга, шунингдек Сизнинг зиммангизга пеня, давлат божи ва вакил хизматлари харажатларини юклашга мажбур бўлади.",
        "Мажбуриятларни виждонан бажаришингизга умид қиламиз ва масалани судгача бўлган тартибда ҳал этишни таклиф қиламиз.",
    ]

    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    safe_contract = re.sub(r"[^A-Za-z0-9_-]+", "_", str(debtor.get("contract_number") or debtor.get("id")))
    pdf_path = GENERATED_DIR / f"claim_uz_{safe_contract}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    image_width = 1654
    image_height = 2339
    margin_left = 90
    margin_right = 90
    margin_top = 70
    margin_bottom = 80
    content_width = image_width - margin_left - margin_right
    center_x = image_width // 2
    line_spacing = 11

    title_font = load_font(30, bold=True)
    section_font = load_font(26, bold=True)
    regular_font = load_font(22)
    small_font = load_font(20)

    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)

    y = margin_top
    centered_header_lines = [
        company_name,
        normalize_text(requisites.get("address")) or "—",
        f"ИНН: {normalize_text(requisites.get('bin')) or '—'}",
    ]
    for line in centered_header_lines:
        y = draw_text_block(
            draw,
            line,
            x=margin_left,
            y=y,
            width=content_width,
            font=section_font if line == company_name else regular_font,
            line_spacing=8,
            align="center",
            paragraph_spacing=4,
        )

    divider_y = y + 10
    draw.line((margin_left, divider_y, image_width - margin_right, divider_y), fill="#111111", width=2)

    right_block_x = image_width - margin_right - 420
    right_y = divider_y + 18
    client_lines = [
        client_name,
        client_address,
        client_phone,
    ]
    for line in client_lines:
        right_y = draw_text_block(
            draw,
            line,
            x=right_block_x,
            y=right_y,
            width=420,
            font=regular_font,
            line_spacing=8,
            align="left",
            paragraph_spacing=2,
        )

    columns_y = right_y + 16
    gutter = 30
    column_width = (content_width - gutter) // 2
    left_x = margin_left
    right_col_x = left_x + column_width + gutter
    vertical_line_x = left_x + column_width + gutter // 2
    draw.line((vertical_line_x, columns_y, vertical_line_x, image_height - margin_bottom - 70), fill="#111111", width=2)

    left_y = columns_y
    right_y = columns_y

    def draw_column_title(x: int, y_value: int, text: str) -> int:
        return draw_text_block(
            draw,
            text,
            x=x,
            y=y_value,
            width=column_width,
            font=section_font,
            line_spacing=8,
            align="center",
            paragraph_spacing=16,
        )

    def draw_column_paragraph(x: int, y_value: int, text: str, *, indent: int = 26, spacing: int = 10) -> int:
        return draw_text_block(
            draw,
            text,
            x=x,
            y=y_value,
            width=column_width,
            font=regular_font,
            line_spacing=line_spacing,
            align="left",
            paragraph_spacing=spacing,
            first_line_indent=indent,
        )

    left_y = draw_column_title(left_x, left_y, "Досудебная претензия")
    right_y = draw_column_title(right_col_x, right_y, "Судгача бўлган талабнома")

    for ru_text, uz_text in zip(ru_paragraphs, uz_paragraphs):
        left_y = draw_column_paragraph(left_x, left_y, ru_text)
        right_y = draw_column_paragraph(right_col_x, right_y, uz_text)

    for product in products:
        display_name = normalize_text(product.get("display_name")) or normalize_text(product.get("name")) or "Товар"
        left_y = draw_column_paragraph(left_x, left_y, display_name, indent=0, spacing=5)
        right_y = draw_column_paragraph(right_col_x, right_y, display_name, indent=0, spacing=5)

    left_y += 12
    right_y += 12
    left_y = draw_column_title(left_x, left_y, "Прошу:")
    right_y = draw_column_title(right_col_x, right_y, "Сўраймиз:")

    left_y = draw_column_paragraph(left_x, left_y, ru_ask[0])
    right_y = draw_column_paragraph(right_col_x, right_y, uz_ask[0])

    for payment_line in payment_detail_lines_ru:
        left_y = draw_column_paragraph(left_x, left_y, payment_line, indent=0, spacing=4)
    for payment_line in payment_detail_lines_uz:
        right_y = draw_column_paragraph(right_col_x, right_y, payment_line, indent=0, spacing=4)

    left_y += 8
    right_y += 8
    for ru_text, uz_text in zip(ru_ask[1:], uz_ask[1:]):
        left_y = draw_column_paragraph(left_x, left_y, ru_text)
        right_y = draw_column_paragraph(right_col_x, right_y, uz_text)

    signature_y = max(left_y, right_y) + 54
    signature_y = min(signature_y, image_height - margin_bottom - 90)
    signature_label_ru = f"Директор {company_name}"
    signature_label_uz = f"Директор {company_name}"
    director_name = normalize_text(requisites.get("director_name")) or "—"
    generated_date = format_date(date.today()) or date.today().isoformat()

    def draw_signature_block(block_x: int, label: str) -> None:
        draw.text((block_x, signature_y), label, fill="#111111", font=small_font)
        draw.text((block_x, signature_y + 34), generated_date, fill="#111111", font=small_font)

        director_width_local = draw.textlength(director_name, font=small_font)
        right_name_x = block_x + max(0, column_width - director_width_local)
        draw.text((right_name_x, signature_y), director_name, fill="#111111", font=small_font)

    draw_signature_block(left_x, signature_label_ru)
    draw_signature_block(right_col_x, signature_label_uz)

    try:
        image.save(pdf_path, "PDF", resolution=150.0)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="РќРµ СѓРґР°Р»РѕСЃСЊ СЃС„РѕСЂРјРёСЂРѕРІР°С‚СЊ PDF РїСЂРµС‚РµРЅР·РёРё.") from exc

    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail="РќРµ СѓРґР°Р»РѕСЃСЊ СЃС„РѕСЂРјРёСЂРѕРІР°С‚СЊ PDF РїСЂРµС‚РµРЅР·РёРё.")

    return pdf_path


def render_claim_pdf(
    debtor: dict[str, Any],
    *,
    debt_amount_override: float | None = None,
    product_overrides: list[dict[str, Any]] | None = None,
) -> Path:
    if normalize_country_code(str(debtor.get("country") or DEFAULT_COUNTRY)) == "uz":
        return _render_claim_pdf_uz(
            debtor,
            debt_amount_override=debt_amount_override,
            product_overrides=product_overrides,
        )
    return _render_claim_pdf_kz(
        debtor,
        debt_amount_override=debt_amount_override,
        product_overrides=product_overrides,
    )


def _render_lawsuit_pdf_old(debtor: dict[str, Any], payload: LawsuitPdfGenerateRequest) -> Path:
    requisites = find_company_requisites(str(debtor.get("company") or ""))
    if requisites is None:
        raise HTTPException(
            status_code=400,
            detail=f"Р”Р»СЏ РєРѕРјРїР°РЅРёРё В«{debtor.get('company') or '—'}В» РЅРµ РЅР°Р№РґРµРЅС‹ СЂРµРєРІРёР·РёС‚С‹.",
        )

    try:
        crm_context = DiSellApiClient().lookup_lawsuit_context(
            str(debtor.get("contract_number") or ""),
            country=str(debtor.get("country") or DEFAULT_COUNTRY),
        )
    except DiSellApiError:
        crm_context = {}

    company_name = requisites.get("company_name") or company_name_value or "—"
    court_name = normalize_text(payload.court_name) or normalize_text(debtor.get("court")) or "—"
    client_name = normalize_text(crm_context.get("client_name")) or normalize_text(debtor.get("client_name")) or "—"
    client_short_name = normalize_text(crm_context.get("client_short_name")) or build_short_client_name(client_name)
    client_inn = normalize_text(crm_context.get("client_inn")) or "—"
    client_address = normalize_text(crm_context.get("client_address")) or normalize_text(debtor.get("address")) or "—"

    crm_phones = crm_context.get("client_phones") or []
    normalized_phones = [normalize_text(phone) for phone in crm_phones if normalize_text(phone)]
    if not normalized_phones:
        normalized_phones = [
            phone
            for phone in [normalize_text(debtor.get("mobile_phone")), normalize_text(debtor.get("home_phone"))]
            if phone
        ]
    client_phones = ", ".join(dict.fromkeys(normalized_phones)) if normalized_phones else "—"

    contract_number = str(crm_context.get("contract_number") or debtor.get("contract_number") or "—")
    contract_date = contract_date_value_from_number(contract_number)
    crm_contract_date = crm_context.get("contract_date")
    if isinstance(crm_contract_date, str) and crm_contract_date:
        try:
            contract_date = date.fromisoformat(crm_contract_date)
        except ValueError:
            pass

    contract_total_amount = parse_float(
        crm_context.get("contract_total_amount")
        if crm_context.get("contract_total_amount") is not None
        else debtor.get("contract_total_amount")
    )
    debt_amount = round(float(payload.debt_amount), 2)
    advance_amount = parse_float(
        crm_context.get("advance_amount")
        if crm_context.get("advance_amount") is not None
        else debtor.get("contract_advance_amount")
    )
    if advance_amount <= 0 and contract_total_amount > 0:
        advance_amount = max(round(contract_total_amount - debt_amount, 2), 0.0)

    discount_amount = parse_float(crm_context.get("discount_amount"))
    installment_balance_amount = max(round(contract_total_amount - advance_amount, 2), 0.0)
    total_payments_amount = parse_float(crm_context.get("advance_amount"))
    post_advance_payments = max(round(total_payments_amount - advance_amount, 2), 0.0)

    product_names = crm_context.get("product_names") or []
    if not product_names:
        product_names = ["РўРѕРІР°СЂ РїРѕ РґРѕРіРѕРІРѕСЂСѓ"]

    penalty_rows, adjusted_penalty_amount, total_overdue_days = build_lawsuit_penalty_rows(
        payload.installment_from,
        payload.installment_to,
        claim_sent_date,
        float(payload.monthly_payment_amount),
        float(payload.first_period_paid_amount or 0),
    )
    penalty_amount = adjusted_penalty_amount
    state_duty_amount = round((debt_amount + penalty_amount) * 0.03, 2)

    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    safe_contract = re.sub(r"[^A-Za-z0-9_-]+", "_", str(debtor.get("contract_number") or debtor.get("id")))
    pdf_path = GENERATED_DIR / f"lawsuit_{safe_contract}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    image_width = 1654
    image_height = 2339
    margin_left = 112
    margin_right = 112
    margin_top = 74
    margin_bottom = 82
    content_width = image_width - margin_left - margin_right
    block_width = 560
    paragraph_indent = 42

    regular_font = load_font(24)
    title_font = load_font(32, bold=True)
    section_font = load_font(24, bold=True)
    small_font = load_font(22)
    table_font = load_font(20)
    table_font_bold = load_font(20, bold=True)

    pages: list[Image.Image] = []
    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)
    current_y = margin_top

    def new_page() -> None:
        nonlocal image, draw, current_y
        pages.append(image)
        image = Image.new("RGB", (image_width, image_height), "white")
        draw = ImageDraw.Draw(image)
        current_y = margin_top

    def estimate_block_height(
        text: str,
        font: ImageFont.ImageFont,
        width: int,
        line_spacing: int,
        first_line_indent: int = 0,
    ) -> int:
        normalized = " ".join((text or "").split())
        if not normalized:
            return 0
        bbox = draw.textbbox((0, 0), "РђРі", font=font)
        line_height = (bbox[3] - bbox[1]) + line_spacing
        if first_line_indent > 0:
            words = normalized.split(" ")
            first_line_words: list[str] = []
            while words:
                candidate_words = first_line_words + [words[0]]
                candidate = " ".join(candidate_words).strip()
                if draw.textlength(candidate, font=font) <= max(40, width - first_line_indent):
                    first_line_words.append(words.pop(0))
                else:
                    break
            remaining = " ".join(words).strip()
            lines = 1 if first_line_words else 0
            if remaining:
                lines += len(wrap_text(draw, remaining, font, width))
            return max(lines, 1) * line_height
        return max(len(wrap_text(draw, normalized, font, width)), 1) * line_height

    def ensure_space(required_height: int) -> None:
        nonlocal current_y
        if current_y + required_height > image_height - margin_bottom:
            new_page()

    def draw_paragraph(
        text: str,
        *,
        font: ImageFont.ImageFont = regular_font,
        align: str = "left",
        spacing_after: int = 12,
        first_line_indent: int = paragraph_indent,
    ) -> None:
        nonlocal current_y
        estimate = estimate_block_height(text, font, content_width, 14, first_line_indent) + spacing_after
        ensure_space(estimate)
        current_y = draw_text_block(
            draw,
            text,
            x=margin_left,
            y=current_y,
            width=content_width,
            font=font,
            line_spacing=14,
            align=align,
            paragraph_spacing=spacing_after,
            first_line_indent=first_line_indent if align == "left" else 0,
        )

    top_right_x = image_width - margin_right - block_width
    current_y = draw_text_block(
        draw,
        court_name,
        x=top_right_x,
        y=current_y,
        width=block_width,
        font=section_font,
        line_spacing=12,
        align="left",
        paragraph_spacing=18,
    )

    plaintiff_lines = [
        "РСЃС‚РµС†:",
        company_name,
        normalize_text(requisites.get("bin")) or "—",
        normalize_text(requisites.get("address")) or "—",
        "Email: sud.process.dp@gmail.com",
        "РўРµР»: +7 700 739 9636",
    ]
    for line in plaintiff_lines:
        current_y = draw_text_block(
            draw,
            line,
            x=top_right_x,
            y=current_y,
            width=block_width,
            font=regular_font if line != "РСЃС‚РµС†:" else section_font,
            line_spacing=10,
            align="left",
            paragraph_spacing=2,
        )

    current_y += 14
    defendant_lines = [
        "РћС‚РІРµС‚С‡РёРє:",
        client_name,
        client_inn,
        client_address,
        client_phones,
    ]
    for line in defendant_lines:
        current_y = draw_text_block(
            draw,
            line,
            x=top_right_x,
            y=current_y,
            width=block_width,
            font=regular_font if line != "РћС‚РІРµС‚С‡РёРє:" else section_font,
            line_spacing=10,
            align="left",
            paragraph_spacing=2,
        )

    current_y += 32
    ensure_space(120)
    current_y = draw_text_block(
        draw,
        "РСЃРє\nРѕ РІР·С‹СЃРєР°РЅРёРё Р·Р°РґРѕР»Р¶РµРЅРЅРѕСЃС‚Рё",
        x=margin_left,
        y=current_y,
        width=content_width,
        font=title_font,
        line_spacing=12,
        align="center",
        paragraph_spacing=28,
    )

    body_paragraphs = [
        "Р”РёСЃРїРѕР·РёС†РёРµР№ СЃС‚Р°С‚СЊРё 9 Р“СЂР°Р¶РґР°РЅСЃРєРѕРіРѕ РєРѕРґРµРєСЃР° Р РµСЃРїСѓР±Р»РёРєРё РљР°Р·Р°С…СЃС‚Р°РЅ (РґР°Р»РµРµ – Р“Рљ) Р·Р°РєСЂРµРїР»РµРЅРѕ, С‡С‚Рѕ Р·Р°С‰РёС‚Р° РіСЂР°Р¶РґР°РЅСЃРєРёС… РїСЂР°РІ РѕСЃСѓС‰РµСЃС‚РІР»СЏРµС‚СЃСЏ СЃСѓРґРѕРј, Р°СЂР±РёС‚СЂР°Р¶РµРј РїСѓС‚РµРј: РїСЂРёР·РЅР°РЅРёСЏ РїСЂР°РІ; РІРѕСЃСЃС‚Р°РЅРѕРІР»РµРЅРёСЏ РїРѕР»РѕР¶РµРЅРёСЏ, СЃСѓС‰РµСЃС‚РІРѕРІР°РІС€РµРіРѕ РґРѕ РЅР°СЂСѓС€РµРЅРёСЏ РїСЂР°РІР°; РїСЂРµСЃРµС‡РµРЅРёСЏ РґРµР№СЃС‚РІРёР№, РЅР°СЂСѓС€Р°СЋС‰РёС… РїСЂР°РІРѕ РёР»Рё СЃРѕР·РґР°СЋС‰РёС… СѓРіСЂРѕР·Сѓ РµРіРѕ РЅР°СЂСѓС€РµРЅРёСЏ; РїСЂРёСЃСѓР¶РґРµРЅРёСЏ Рє РёСЃРїРѕР»РЅРµРЅРёСЋ РѕР±СЏР·Р°РЅРЅРѕСЃС‚Рё РІ РЅР°С‚СѓСЂРµ; РІР·С‹СЃРєР°РЅРёСЏ СѓР±С‹С‚РєРѕРІ, РЅРµСѓСЃС‚РѕР№РєРё; РїСЂРёР·РЅР°РЅРёСЏ РѕСЃРїРѕСЂРёРјРѕР№ СЃРґРµР»РєРё РЅРµРґРµР№СЃС‚РІРёС‚РµР»СЊРЅРѕР№ и применения последствий ее недействительности, применения последствий недействительности ничтожной сделки; компенсации морального вреда; прекращения или изменения правоотношений; признания недействительным или не подлежащим применению не соответствующего законодательству Республики Казахстан акта органа государственного управления или местного представительного либо исполнительного органа; взыскания штрафа с государственного органа или должностного лица за воспрепятствование гражданину или юридическому лицу в приобретении или осуществлении права, а также иными способами, предусмотренными законодательными актами.",
        f"РњРµР¶РґСѓ {company_name} и {client_short_name} заключен договор купли/продажи № {contract_number} от {format_date(contract_date) or '—'}, в соответствии с которым в собственность передан товар, а именно:",
    ]
    for paragraph in body_paragraphs:
        draw_paragraph(paragraph)

    for product_name in product_names:
        draw_paragraph(product_name, first_line_indent=0, spacing_after=6)

    remaining_paragraphs = [
        "РЎС‚Р°С‚СЊРµР№ 7 ГК определено, что гражданские права и обязанности возникают, изменяются и прекращаются из оснований, предусмотренных законодательством Республики Казахстан, а также из действий граждан и юридических лиц, которые хотя и не предусмотрены им, но в силу общих начал и смысла гражданского законодательства порождают гражданские права и обязанности. В соответствии с этим гражданские права и обязанности возникают, изменяются и прекращаются, помимо прочих, из договоров и иных сделок, предусмотренных законодательством Республики Казахстан, а также из сделок, хотя и не предусмотренных им, но не противоречащих законодательству Республики Казахстан.",
        "РЎС‚Р°С‚СЊРµР№ 378 ГК определено, что договором признается соглашение двух или нескольких лиц об установлении, изменении или прекращении гражданских прав и обязанностей.",
        "В силу статьи 393 ГК договор считается заключенным, когда между сторонами в требуемой в подлежащих случаях форме достигнуто соглашение по всем существенным его условиям. Существенными являются условия о предмете договора, условия, которые признаны существенными законодательством или необходимы для договоров данного вида, а также все те условия, относительно которых по заявлению одной из сторон должно быть достигнуто соглашение. Если в соответствии с законодательными актами для заключения договора необходима передача имущества, договор считается заключенным с момента передачи соответствующего имущества.",
        "В соответствии со статьей 406 ГК по договору купли-продажи одна сторона (продавец) обязуется передать имущество (товар) в собственность, хозяйственное ведение или оперативное управление другой стороне (покупателю), а покупатель обязуется принять это имущество (товар) и уплатить за него определенную денежную сумму (цену).",
        f"Согласно условиям заключенного договора определено, что ответчик принял на себя обязательства по оплате стоимости полученного товара. Стоимость переданного товара с учетом скидки {format_money(discount_amount)} тенге была определена в размере {format_money(contract_total_amount)} тенге, из которых {format_money(advance_amount)} тенге были оплачены в качестве предоплаты, оставшаяся сумма в размере {format_money(installment_balance_amount)} тенге подлежала оплате в рассрочку в период с {format_date(payload.installment_from) or '—'} года по {format_date(payload.installment_to) or '—'} года равными платежами по {format_money(payload.monthly_payment_amount)} тенге.",
        f"Ответчиком были внесены платежи в размере {format_money(post_advance_payments)} тенге, в связи с чем сумма задолженности на данный момент составила {format_money(debt_amount)} тенге.",
        f"Однако обязательства, принятые на себя, ответчик на сумму в размере {format_money(debt_amount)} тенге не исполнил, что привело к нарушению прав и охраняемых законом интересов {company_name}.",
        "В силу статьи 272 ГК обязательство должно исполняться надлежащим образом в соответствии с условиями обязательства и требованиями законодательства, а при отсутствии таких условий и требований - в соответствии с обычаями делового оборота или иными обычно предъявляемыми требованиями. Согласно статье 277 ГК, если обязательство предусматривает или позволяет определить день его исполнения или период времени, в течение которого оно должно быть исполнено, обязательство подлежит исполнению в этот день или, соответственно, в любой момент в пределах такого периода.",
        "В соответствии с пунктом 1 статьи 353 Гражданского кодекса Республики Казахстан, при неисполнении или ненадлежащем исполнении денежного обязательства должник обязан уплатить кредитору неустойку (пеню) в размере 0,1% от суммы долга за каждый день просрочки.",
        f"На основании вышеуказанного положения закона, произведён расчёт пени по обязательству в размере {format_money(debt_amount)} тенге, {total_overdue_days} календарных дней просрочки, с учётом поэтапного увеличения задолженности следующим образом: {format_money(penalty_amount)} тенге, согласно нижеприведенному расчету:",
    ]
    for paragraph in remaining_paragraphs:
        draw_paragraph(paragraph)

    table_header_height = 76
    row_height = 44
    table_total_height = table_header_height + (len(penalty_rows) * row_height) + row_height + 42
    ensure_space(table_total_height)
    table_x = margin_left
    table_y = current_y + 8
    col_widths = [150, 150, 180, 360, 160, 210]

    def draw_cell(x: int, y: int, width: int, height: int, text: str, *, font: ImageFont.ImageFont, align: str = "center", bold: bool = False) -> None:
        draw.rectangle((x, y, x + width, y + height), outline="#111111", width=1)
        text_font = table_font_bold if bold else font
        lines = wrap_text(draw, text, text_font, max(width - 10, 20))
        bbox = draw.textbbox((0, 0), "РђРі", font=text_font)
        line_height = (bbox[3] - bbox[1]) + 4
        block_height = len(lines) * line_height
        line_y = y + max((height - block_height) / 2, 4)
        for line in lines:
            line_width = draw.textlength(line, font=text_font)
            if align == "left":
                line_x = x + 6
            elif align == "right":
                line_x = x + width - line_width - 6
            else:
                line_x = x + max((width - line_width) / 2, 4)
            draw.text((line_x, line_y), line, fill="#111111", font=text_font)
            line_y += line_height

    first_header_y = table_y
    second_header_y = table_y + int(table_header_height / 2)
    x = table_x
    draw_cell(x, first_header_y, col_widths[0] + col_widths[1], int(table_header_height / 2), "РџРµСЂРёРѕРґ", font=table_font, bold=True)
    x += col_widths[0] + col_widths[1]
    for label, width in zip(
        [
            "РљРѕР»РёС‡РµСЃС‚РІРѕ РґРЅРµР№ РїСЂРѕСЃСЂРѕС‡РєРё",
            "РЎСѓРјРјР° РЅРµРёСЃРїРѕР»РЅРµРЅРЅРѕРіРѕ РѕР±СЏР·Р°С‚РµР»СЊСЃС‚РІР° (С‚РµРЅРіРµ)",
            "Р Р°Р·РјРµСЂ РїРµРЅРё (РґРµРЅСЊ %)",
            "РЎСѓРјРјР° РїРµРЅРё (С‚РµРЅРіРµ)",
        ],
        col_widths[2:],
    ):
        draw_cell(x, first_header_y, width, table_header_height, label, font=table_font, bold=True)
        x += width
    draw_cell(table_x, second_header_y, col_widths[0], int(table_header_height / 2), "РћС‚", font=table_font, bold=True)
    draw_cell(table_x + col_widths[0], second_header_y, col_widths[1], int(table_header_height / 2), "Р”Рѕ", font=table_font, bold=True)

    current_row_y = table_y + table_header_height
    for row in penalty_rows:
        cell_values = [
            format_date(row["period_from"]) or "—",
            format_date(row["period_to"]) or "—",
            str(row["days"]),
            format_money(row["obligation_amount"]),
            "0,1",
            format_money(row["penalty_amount"]),
        ]
        x = table_x
        for value, width in zip(cell_values, col_widths):
            draw_cell(x, current_row_y, width, row_height, value, font=table_font)
            x += width
        current_row_y += row_height

    summary_width = sum(col_widths[:-1])
    draw_cell(table_x, current_row_y, summary_width, row_height, "РС‚РѕРіРѕ", font=table_font, align="right", bold=True)
    draw_cell(table_x + summary_width, current_row_y, col_widths[-1], row_height, format_money(penalty_amount), font=table_font, bold=True)
    current_y = current_row_y + row_height + 20

    closing_paragraphs = [
        f"Требования {company_name}, адресованные {client_short_name}, о необходимости исполнения обязательства оставлены последним без удовлетворения.",
        f"Совокупность приведенных норм законодательства и изложенных обстоятельств позволяет сделать вывод о том, что с {client_short_name} в пользу {company_name} подлежат взысканию сумма задолженности в размере {format_money(debt_amount)} тенге и пеня в размере {format_money(penalty_amount)} тенге.",
        "В силу статьи 4 Гражданского процессуального кодекса Республики Казахстан (далее – ГПК) задачами гражданского судопроизводства являются защита и восстановление нарушенных или оспариваемых прав, свобод и законных интересов граждан, государства и юридических лиц, соблюдение законности в гражданском обороте, обеспечение полного и своевременного рассмотрения дела, содействие мирному урегулированию спора, предупреждение правонарушений и формирование в обществе уважительного отношения к закону и суду.",
        "На основании изложенного, руководствуясь статьями 148, 149, ГПК РК",
    ]
    for paragraph in closing_paragraphs:
        draw_paragraph(paragraph)

    ensure_space(260)
    current_y = draw_text_block(
        draw,
        "РџСЂРѕС€Сѓ:",
        x=margin_left,
        y=current_y + 6,
        width=content_width,
        font=section_font,
        line_spacing=12,
        align="left",
        paragraph_spacing=18,
    )

    ask_line = (
        f"1. Взыскать с {client_name} в пользу {company_name} сумму задолженности в размере "
        f"{format_money(debt_amount)} ({money_to_words_ru(debt_amount)}) тенге, пеню в размере "
        f"{format_money(penalty_amount)} ({money_to_words_ru(penalty_amount)}) тенге, государственную пошлину "
        f"в размере {format_money(state_duty_amount)} ({money_to_words_ru(state_duty_amount)}) тенге."
    )
    draw_paragraph(ask_line, first_line_indent=0, spacing_after=18)

    current_y = draw_text_block(
        draw,
        "РџСЂРёР»РѕР¶РµРЅРёРµ:",
        x=margin_left,
        y=current_y,
        width=content_width,
        font=section_font,
        line_spacing=12,
        align="left",
        paragraph_spacing=10,
    )

    attachments = [
        "1) Квитанция об уплате государственной пошлины;",
        "2) Копия договора купли-продажи с актом приема-передачи;",
        "3) Копия устава;",
        "4) Копия свидетельства (справка о государственной регистрации);",
        "5) Копия досудебной претензии с квитанцией об отправке.",
    ]
    for item in attachments:
        draw_paragraph(item, first_line_indent=0, spacing_after=6)

    signature_y = max(current_y + 24, image_height - margin_bottom - 84)
    director_name = normalize_text(requisites.get("director_name")) or "—"
    left_signature = f"Директор {company_name}"
    draw.text((margin_left, signature_y), left_signature, fill="#111111", font=small_font)
    director_width = draw.textlength(director_name, font=small_font)
    draw.text((image_width - margin_right - director_width, signature_y), director_name, fill="#111111", font=small_font)
    generated_date = format_date(date.today()) or date.today().isoformat()
    draw.text((margin_left, signature_y + 38), generated_date, fill="#111111", font=small_font)

    pages.append(image)
    try:
        first_page, *other_pages = pages
        first_page.save(pdf_path, "PDF", resolution=150.0, save_all=True, append_images=other_pages)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF иска.") from exc

    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF иска.")

    return pdf_path

def _render_lawsuit_pdf_uz(debtor: dict[str, Any], payload: LawsuitPdfGenerateRequest) -> Path:
    company_name_value = sanitize_company_name(str(debtor.get("company") or ""))
    requisites = find_company_requisites(company_name_value)
    if requisites is None:
        raise HTTPException(
            status_code=400,
            detail=f"Для компании «{company_name_value or '—'}» не найдены реквизиты.",
        )

    claim_sent_date_raw = debtor.get("claim_sent_date")
    if not claim_sent_date_raw:
        raise HTTPException(status_code=400, detail="Для генерации иска сначала заполните дату отправки претензии.")
    try:
        claim_sent_date = date.fromisoformat(str(claim_sent_date_raw))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Дата отправки претензии заполнена некорректно.") from exc

    try:
        crm_context = DiSellApiClient().lookup_lawsuit_context(
            str(debtor.get("contract_number") or ""),
            country="uz",
        )
    except DiSellApiError:
        crm_context = {}

    company_name = requisites.get("company_name") or company_name_value or "—"
    court_name = normalize_text(payload.court_name) or normalize_text(debtor.get("court")) or "—"
    client_name = normalize_text(crm_context.get("client_name")) or normalize_text(debtor.get("client_name")) or "—"
    client_short_name = normalize_text(crm_context.get("client_short_name")) or build_short_client_name(client_name)
    client_inn = normalize_text(crm_context.get("client_inn")) or "—"
    client_address = normalize_text(crm_context.get("client_address")) or normalize_text(debtor.get("address")) or "—"

    crm_phones = crm_context.get("client_phones") or []
    normalized_phones = [normalize_text(phone) for phone in crm_phones if normalize_text(phone)]
    if not normalized_phones:
        normalized_phones = [
            phone
            for phone in [normalize_text(debtor.get("mobile_phone")), normalize_text(debtor.get("home_phone"))]
            if phone
        ]
    client_phones = ", ".join(dict.fromkeys(normalized_phones)) if normalized_phones else "—"

    contract_number = str(crm_context.get("contract_number") or debtor.get("contract_number") or "—")
    contract_date = contract_date_value_from_number(contract_number)
    crm_contract_date = crm_context.get("contract_date")
    if isinstance(crm_contract_date, str) and crm_contract_date:
        try:
            contract_date = date.fromisoformat(crm_contract_date)
        except ValueError:
            pass

    contract_total_amount = parse_float(
        crm_context.get("contract_total_amount")
        if crm_context.get("contract_total_amount") is not None
        else debtor.get("contract_total_amount")
    )
    debt_amount = parse_float(
        crm_context.get("debt_amount")
        if crm_context.get("debt_amount") is not None
        else payload.debt_amount
    )
    advance_amount = parse_float(
        crm_context.get("advance_amount")
        if crm_context.get("advance_amount") is not None
        else debtor.get("contract_advance_amount")
    )
    if advance_amount <= 0 and contract_total_amount > 0:
        advance_amount = max(round(contract_total_amount - debt_amount, 2), 0.0)

    discount_amount = parse_float(crm_context.get("discount_amount"))
    installment_balance_amount = max(round(contract_total_amount - advance_amount, 2), 0.0)
    products = normalize_document_products(
        [item.model_dump() for item in (payload.product_overrides or [])] or crm_context.get("products"),
        fallback_name="Товар по договору",
    )

    last_missed_payment_date = None
    last_missed_raw = debtor.get("last_missed_payment_date")
    if isinstance(last_missed_raw, str) and last_missed_raw:
        try:
            last_missed_payment_date = date.fromisoformat(last_missed_raw)
        except ValueError:
            last_missed_payment_date = None
    total_overdue_days = max((claim_sent_date - last_missed_payment_date).days, 0) if last_missed_payment_date else int(debtor.get("debt_days") or 0)
    penalty_amount = compute_simple_penalty_amount(debt_amount, total_overdue_days)
    claim_price_amount = compute_lawsuit_claim_price(debt_amount, penalty_amount)
    state_duty_amount = compute_lawsuit_state_duty("uz", debt_amount, penalty_amount)

    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    safe_contract = re.sub(r"[^A-Za-z0-9_-]+", "_", str(debtor.get("contract_number") or debtor.get("id")))
    pdf_path = GENERATED_DIR / f"lawsuit_uz_{safe_contract}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    image_width = 1654
    image_height = 2339
    margin_left = 112
    margin_right = 112
    margin_top = 72
    margin_bottom = 82
    content_width = image_width - margin_left - margin_right
    block_gap = 20
    half_width = content_width
    paragraph_indent = 42

    regular_font = load_font(24)
    title_font = load_font(31, bold=True)
    section_font = load_font(24, bold=True)
    small_font = load_font(21)
    table_font = load_font(19)
    table_font_bold = load_font(19, bold=True)

    pages: list[Image.Image] = []
    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)
    current_y = margin_top

    def new_page() -> None:
        nonlocal image, draw, current_y
        pages.append(image)
        image = Image.new("RGB", (image_width, image_height), "white")
        draw = ImageDraw.Draw(image)
        current_y = margin_top

    def estimate_block_height(text: str, font: ImageFont.ImageFont, width: int, line_spacing: int, first_line_indent: int = 0) -> int:
        normalized = " ".join((text or "").split())
        if not normalized:
            return 0
        bbox = draw.textbbox((0, 0), "Аг", font=font)
        line_height = (bbox[3] - bbox[1]) + line_spacing
        if first_line_indent > 0:
            words = normalized.split(" ")
            first_line_words: list[str] = []
            while words:
                candidate_words = first_line_words + [words[0]]
                candidate = " ".join(candidate_words).strip()
                if draw.textlength(candidate, font=font) <= max(40, width - first_line_indent):
                    first_line_words.append(words.pop(0))
                else:
                    break
            remaining = " ".join(words).strip()
            lines = 1 if first_line_words else 0
            if remaining:
                lines += len(wrap_text(draw, remaining, font, width))
            return max(lines, 1) * line_height
        return max(len(wrap_text(draw, normalized, font, width)), 1) * line_height

    def ensure_space(required_height: int) -> None:
        nonlocal current_y
        if current_y + required_height > image_height - margin_bottom:
            new_page()

    def draw_paragraph(
        text: str,
        *,
        font: ImageFont.ImageFont = regular_font,
        align: str = "left",
        spacing_after: int = 12,
        first_line_indent: int = paragraph_indent,
        width: int = content_width,
        x: int = margin_left,
    ) -> None:
        nonlocal current_y
        estimate = estimate_block_height(text, font, width, 12, first_line_indent) + spacing_after
        ensure_space(estimate)
        current_y = draw_text_block(
            draw,
            text,
            x=x,
            y=current_y,
            width=width,
            font=font,
            line_spacing=12,
            align=align,
            paragraph_spacing=spacing_after,
            first_line_indent=first_line_indent if align == "left" else 0,
        )

    def draw_block(lines: list[str], x: int, y: int, width: int, title: str) -> int:
        block_y = draw_text_block(
            draw,
            title,
            x=x,
            y=y,
            width=width,
            font=section_font,
            line_spacing=10,
            align="left",
            paragraph_spacing=6,
        )
        for line in lines:
            block_y = draw_text_block(
                draw,
                line,
                x=x,
                y=block_y,
                width=width,
                font=regular_font,
                line_spacing=8,
                align="left",
                paragraph_spacing=2,
            )
        return block_y

    court_lines = [line.strip() for line in str(court_name).splitlines() if line.strip()] or [court_name]
    for line in court_lines:
        current_y = draw_text_block(
            draw,
            line,
            x=margin_left,
            y=current_y,
            width=content_width,
            font=section_font,
            line_spacing=10,
            align="left",
            paragraph_spacing=2,
        )
    current_y += 18

    left_block_lines = [
        company_name,
        normalize_text(requisites.get("address")) or "—",
        f"ИНН: {normalize_text(requisites.get('bin')) or '—'}",
        "Тел.:",
        f"МФО: {normalize_text(requisites.get('bank_mfo')) or '—'}",
    ]
    right_block_lines = [
        client_name,
        f"ИНН: {client_inn}",
        client_address,
        f"Тел.: {client_phones}",
        f"Цена иска: {format_money(claim_price_amount)} сум",
    ]
    block_top = current_y
    plaintiff_bottom = draw_block(left_block_lines, margin_left, block_top, half_width, "Истец:")
    defendant_top = plaintiff_bottom + block_gap
    defendant_bottom = draw_block(
        right_block_lines,
        margin_left,
        defendant_top,
        half_width,
        "Ответчик:",
    )
    current_y = defendant_bottom + 28

    current_y = draw_text_block(
        draw,
        "Иск\nо взыскании задолженности",
        x=margin_left,
        y=current_y,
        width=content_width,
        font=title_font,
        line_spacing=10,
        align="center",
        paragraph_spacing=24,
    )

    contract_amount_sentence = (
        f"Согласно условиям заключенного договора определено, что ответчик принял на себя обязательства по оплате стоимости полученного товара. "
        f"Стоимость переданного товара была определена в размере {format_money(contract_total_amount)} сум, из которых "
        f"{format_money(advance_amount)} сум были оплачены в качестве предоплаты, оставшаяся сумма в размере "
        f"{format_money(installment_balance_amount)} сум подлежала оплате в соответствии с графиком платежей, предусмотренным договором."
    )
    if discount_amount > 0:
        contract_amount_sentence = (
            f"Согласно условиям заключенного договора определено, что ответчик принял на себя обязательства по оплате стоимости полученного товара. "
            f"Стоимость переданного товара с учетом скидки {format_money(discount_amount)} сум была определена в размере "
            f"{format_money(contract_total_amount)} сум, из которых {format_money(advance_amount)} сум были оплачены в качестве предоплаты, "
            f"оставшаяся сумма в размере {format_money(installment_balance_amount)} сум подлежала оплате в соответствии с графиком платежей, предусмотренным договором."
        )

    intro_paragraphs = [
        "В соответствии со статьями 353, 357 и 364 Гражданского кодекса Республики Узбекистан обязательства возникают из договора и подлежат обязательному исполнению сторонами в соответствии с его условиями.",
        f"Между {company_name} и {client_short_name} заключен договор купли-продажи № {contract_number} от {format_date(contract_date) or '—'}, в соответствии с которым в собственность ответчика был передан товар, а именно:",
    ]
    for paragraph in intro_paragraphs:
        draw_paragraph(paragraph)

    for product in products:
        draw_paragraph(
            normalize_text(product.get('display_name')) or normalize_text(product.get("name")) or "Товар",
            first_line_indent=0,
            spacing_after=6,
        )

    body_paragraphs = [
        "Статьей 386 Гражданского кодекса Республики Узбекистан предусмотрено, что по договору купли-продажи одна сторона обязуется передать товар в собственность другой стороне, а покупатель — принять товар и уплатить за него установленную договором денежную сумму.",
        "В силу статьи 236 Гражданского кодекса Республики Узбекистан обязательства должны исполняться надлежащим образом в соответствии с условиями обязательства и требованиями законодательства.",
        contract_amount_sentence,
        f"Ответчиком были внесены платежи в размере {format_money(advance_amount)} сум, в связи с чем сумма задолженности на текущий момент составляет {format_money(debt_amount)} сум.",
        f"Однако обязательства, принятые на себя, ответчик на сумму в размере {format_money(debt_amount)} сум не исполнил, что привело к нарушению прав и охраняемых законом интересов {company_name}.",
        "Пунктом 2.3 заключенного договора определено, что в случае просрочки очередного платежа при оплате товара в рассрочку покупатель уплачивает продавцу пеню в размере 0,1 % за каждый день просрочки от суммы платежа, подлежащего уплате.",
        f"На основании указанных условий произведен расчет пени по обязательству в размере {format_money(debt_amount)} сум за {total_overdue_days} календарных дней просрочки, что составляет {format_money(penalty_amount)} сум.",
    ]
    for paragraph in body_paragraphs:
        draw_paragraph(paragraph)

    closing_paragraphs = [
        f"Требования {company_name}, адресованные {client_short_name}, о необходимости исполнения обязательства оставлены последним без удовлетворения.",
        f"Совокупность приведенных норм законодательства и изложенных обстоятельств позволяет сделать вывод о том, что с {client_short_name} в пользу {company_name} подлежат взысканию сумма задолженности в размере {format_money(debt_amount)} сум и пеня в размере {format_money(penalty_amount)} сум.",
        f"Цена иска составляет {format_money(claim_price_amount)} сум. Государственная пошлина по настоящему иску составляет 4 процента от цены иска, но не менее 1 БРВ, и в данном случае равна {format_money(state_duty_amount)} сум.",
        "В силу статьи 4 Гражданского процессуального кодекса Республики Узбекистан задачами гражданского судопроизводства являются защита и восстановление нарушенных прав и законных интересов граждан и юридических лиц, обеспечение законности и своевременного рассмотрения дела.",
        "На основании изложенного, руководствуясь статьями 4, 193, 257 Гражданского процессуального кодекса Республики Узбекистан,",
    ]
    for paragraph in closing_paragraphs:
        draw_paragraph(paragraph)

    current_y = draw_text_block(
        draw,
        "Прошу:",
        x=margin_left,
        y=current_y + 6,
        width=content_width,
        font=section_font,
        line_spacing=10,
        align="left",
        paragraph_spacing=18,
    )

    ask_line = (
        f"1. Взыскать с {client_name} в пользу {company_name} сумму задолженности в размере "
        f"{format_money(debt_amount)} ({money_to_words_sum_ru(debt_amount)}) сум, пеню в размере "
        f"{format_money(penalty_amount)} ({money_to_words_sum_ru(penalty_amount)}) сум, государственную пошлину "
        f"в размере {format_money(state_duty_amount)} ({money_to_words_sum_ru(state_duty_amount)}) сум."
    )
    draw_paragraph(ask_line, first_line_indent=0, spacing_after=18)

    current_y = draw_text_block(
        draw,
        "Приложения:",
        x=margin_left,
        y=current_y,
        width=content_width,
        font=section_font,
        line_spacing=10,
        align="left",
        paragraph_spacing=10,
    )

    attachments = [
        "1) Квитанция об уплате государственной пошлины;",
        "2) Копия договора купли-продажи с актом приема-передачи;",
        "3) Копия устава;",
        "4) Копия свидетельства (справка о государственной регистрации);",
        "5) Копия досудебной претензии с квитанцией об отправке.",
    ]
    for item in attachments:
        draw_paragraph(item, first_line_indent=0, spacing_after=6)

    signature_y = max(current_y + 24, image_height - margin_bottom - 84)
    director_name = normalize_text(requisites.get("director_name")) or "—"
    left_signature = f"Директор {company_name}"
    draw.text((margin_left, signature_y), left_signature, fill="#111111", font=small_font)
    director_width = draw.textlength(director_name, font=small_font)
    draw.text((image_width - margin_right - director_width, signature_y), director_name, fill="#111111", font=small_font)
    generated_date = format_date(date.today()) or date.today().isoformat()
    draw.text((margin_left, signature_y + 34), generated_date, fill="#111111", font=small_font)

    pages.append(image)
    try:
        first_page, *other_pages = pages
        first_page.save(pdf_path, "PDF", resolution=150.0, save_all=True, append_images=other_pages)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF иска.") from exc

    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF иска.")

    return pdf_path


def render_lawsuit_pdf(debtor: dict[str, Any], payload: LawsuitPdfGenerateRequest) -> Path:
    country_code = normalize_country_code(str(debtor.get("country") or DEFAULT_COUNTRY))
    if country_code == "uz":
        return _render_lawsuit_pdf_uz(debtor, payload)

    company_name_value = sanitize_company_name(str(debtor.get("company") or ""))
    requisites = find_company_requisites(company_name_value)
    if requisites is None:
        raise HTTPException(
            status_code=400,
            detail=f"Для компании «{company_name_value or '—'}» не найдены реквизиты.",
        )

    claim_sent_date_raw = debtor.get("claim_sent_date")
    if not claim_sent_date_raw:
        raise HTTPException(status_code=400, detail="Для генерации иска сначала заполните дату отправки претензии.")
    try:
        claim_sent_date = date.fromisoformat(str(claim_sent_date_raw))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Дата отправки претензии заполнена некорректно.") from exc

    try:
        crm_context = DiSellApiClient().lookup_lawsuit_context(
            str(debtor.get("contract_number") or ""),
            country=str(debtor.get("country") or DEFAULT_COUNTRY),
        )
    except DiSellApiError:
        crm_context = {}

    company_name = requisites.get("company_name") or company_name_value or "—"
    court_name = normalize_text(payload.court_name) or normalize_text(debtor.get("court")) or "—"
    client_name = normalize_text(crm_context.get("client_name")) or normalize_text(debtor.get("client_name")) or "—"
    client_short_name = normalize_text(crm_context.get("client_short_name")) or build_short_client_name(client_name)
    client_inn = normalize_text(crm_context.get("client_inn")) or "—"
    client_address = normalize_text(crm_context.get("client_address")) or normalize_text(debtor.get("address")) or "—"

    crm_phones = crm_context.get("client_phones") or []
    normalized_phones = [normalize_text(phone) for phone in crm_phones if normalize_text(phone)]
    if not normalized_phones:
        normalized_phones = [
            phone
            for phone in [normalize_text(debtor.get("mobile_phone")), normalize_text(debtor.get("home_phone"))]
            if phone
        ]
    client_phones = ", ".join(dict.fromkeys(normalized_phones)) if normalized_phones else "—"

    contract_number = str(crm_context.get("contract_number") or debtor.get("contract_number") or "—")
    contract_date = contract_date_value_from_number(contract_number)
    crm_contract_date = crm_context.get("contract_date")
    if isinstance(crm_contract_date, str) and crm_contract_date:
        try:
            contract_date = date.fromisoformat(crm_contract_date)
        except ValueError:
            pass

    contract_total_amount = parse_float(
        crm_context.get("contract_total_amount")
        if crm_context.get("contract_total_amount") is not None
        else debtor.get("contract_total_amount")
    )
    debt_amount = round(float(payload.debt_amount), 2)
    advance_amount = parse_float(
        crm_context.get("advance_amount")
        if crm_context.get("advance_amount") is not None
        else debtor.get("contract_advance_amount")
    )
    if advance_amount <= 0 and contract_total_amount > 0:
        advance_amount = max(round(contract_total_amount - debt_amount, 2), 0.0)

    discount_amount = parse_float(crm_context.get("discount_amount"))
    installment_balance_amount = max(round(contract_total_amount - advance_amount, 2), 0.0)
    total_payments_amount = parse_float(crm_context.get("advance_amount"))
    post_advance_payments = max(round(total_payments_amount - advance_amount, 2), 0.0)

    products = normalize_document_products(
        [item.model_dump() for item in (payload.product_overrides or [])] or crm_context.get("products"),
        fallback_name="Товар по договору",
    )

    penalty_rows, adjusted_penalty_amount, total_overdue_days = build_lawsuit_penalty_rows(
        payload.installment_from,
        payload.installment_to,
        claim_sent_date,
        float(payload.monthly_payment_amount),
        float(payload.first_period_paid_amount or 0),
    )
    penalty_amount = adjusted_penalty_amount
    state_duty_amount = compute_lawsuit_state_duty(country_code, debt_amount, penalty_amount)

    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    safe_contract = re.sub(r"[^A-Za-z0-9_-]+", "_", str(debtor.get("contract_number") or debtor.get("id")))
    pdf_path = GENERATED_DIR / f"lawsuit_{safe_contract}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    image_width = 1654
    image_height = 2339
    margin_left = 112
    margin_right = 112
    margin_top = 74
    margin_bottom = 82
    content_width = image_width - margin_left - margin_right
    block_width = 560
    paragraph_indent = 42

    regular_font = load_font(24)
    title_font = load_font(32, bold=True)
    section_font = load_font(24, bold=True)
    small_font = load_font(22)
    table_font = load_font(20)
    table_font_bold = load_font(20, bold=True)

    pages: list[Image.Image] = []
    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)
    current_y = margin_top

    def new_page() -> None:
        nonlocal image, draw, current_y
        pages.append(image)
        image = Image.new("RGB", (image_width, image_height), "white")
        draw = ImageDraw.Draw(image)
        current_y = margin_top

    def estimate_block_height(text: str, font: ImageFont.ImageFont, width: int, line_spacing: int, first_line_indent: int = 0) -> int:
        normalized = " ".join((text or "").split())
        if not normalized:
            return 0
        bbox = draw.textbbox((0, 0), "Аг", font=font)
        line_height = (bbox[3] - bbox[1]) + line_spacing
        if first_line_indent > 0:
            words = normalized.split(" ")
            first_line_words: list[str] = []
            while words:
                candidate_words = first_line_words + [words[0]]
                candidate = " ".join(candidate_words).strip()
                if draw.textlength(candidate, font=font) <= max(40, width - first_line_indent):
                    first_line_words.append(words.pop(0))
                else:
                    break
            remaining = " ".join(words).strip()
            lines = 1 if first_line_words else 0
            if remaining:
                lines += len(wrap_text(draw, remaining, font, width))
            return max(lines, 1) * line_height
        return max(len(wrap_text(draw, normalized, font, width)), 1) * line_height

    def ensure_space(required_height: int) -> None:
        nonlocal current_y
        if current_y + required_height > image_height - margin_bottom:
            new_page()

    def draw_paragraph(
        text: str,
        *,
        font: ImageFont.ImageFont = regular_font,
        align: str = "left",
        spacing_after: int = 12,
        first_line_indent: int = paragraph_indent,
    ) -> None:
        nonlocal current_y
        estimate = estimate_block_height(text, font, content_width, 14, first_line_indent) + spacing_after
        ensure_space(estimate)
        current_y = draw_text_block(
            draw,
            text,
            x=margin_left,
            y=current_y,
            width=content_width,
            font=font,
            line_spacing=14,
            align=align,
            paragraph_spacing=spacing_after,
            first_line_indent=first_line_indent if align == "left" else 0,
        )

    top_right_x = image_width - margin_right - block_width
    current_y = draw_text_block(
        draw,
        court_name,
        x=top_right_x,
        y=current_y,
        width=block_width,
        font=section_font,
        line_spacing=12,
        align="left",
        paragraph_spacing=18,
    )

    plaintiff_lines = [
        "Истец:",
        company_name,
        f"БИН: {normalize_text(requisites.get('bin')) or '—'}",
        normalize_text(requisites.get("address")) or "—",
        "Email: sud.process.dp@gmail.com",
        "Тел: +7 700 739 9636",
    ]
    for line in plaintiff_lines:
        current_y = draw_text_block(
            draw,
            line,
            x=top_right_x,
            y=current_y,
            width=block_width,
            font=regular_font if line != "Истец:" else section_font,
            line_spacing=10,
            align="left",
            paragraph_spacing=2,
        )

    current_y += 14
    defendant_lines = [
        "Ответчик:",
        client_name,
        f"ИИН: {client_inn}",
        client_address,
        client_phones,
    ]
    for line in defendant_lines:
        current_y = draw_text_block(
            draw,
            line,
            x=top_right_x,
            y=current_y,
            width=block_width,
            font=regular_font if line != "Ответчик:" else section_font,
            line_spacing=10,
            align="left",
            paragraph_spacing=2,
        )

    current_y += 32
    ensure_space(120)
    current_y = draw_text_block(
        draw,
        "Иск\nо взыскании задолженности",
        x=margin_left,
        y=current_y,
        width=content_width,
        font=title_font,
        line_spacing=12,
        align="center",
        paragraph_spacing=28,
    )

    intro_paragraphs = [
        "Диспозицией статьи 9 Гражданского кодекса Республики Казахстан (далее – ГК) закреплено, что защита гражданских прав осуществляется судом, арбитражем путем: признания прав; восстановления положения, существовавшего до нарушения права; пресечения действий, нарушающих право или создающих угрозу его нарушения; присуждения к исполнению обязанности в натуре; взыскания убытков, неустойки; признания оспоримой сделки недействительной и применения последствий ее недействительности, применения последствий недействительности ничтожной сделки; компенсации морального вреда; прекращения или изменения правоотношений; признания недействительным или не подлежащим применению не соответствующего законодательству Республики Казахстан акта органа государственного управления или местного представительного либо исполнительного органа; взыскания штрафа с государственного органа или должностного лица за воспрепятствование гражданину или юридическому лицу в приобретении или осуществлении права, а также иными способами, предусмотренными законодательными актами Республики Казахстан.",
        f"Между {company_name} и {client_short_name} заключен договор купли/продажи № {contract_number} от {format_date(contract_date) or '—'}, в соответствии с которым в собственность передан товар, а именно:",
    ]
    for paragraph in intro_paragraphs:
        draw_paragraph(paragraph)

    for product in products:
        draw_paragraph(
            normalize_text(product.get("display_name")) or normalize_text(product.get("name")) or "Товар",
            first_line_indent=0,
            spacing_after=6,
        )

    main_paragraphs = [
        "Статьей 7 ГК определено, что гражданские права и обязанности возникают, изменяются и прекращаются из оснований, предусмотренных законодательством Республики Казахстан, а также из действий граждан и юридических лиц, которые хотя и не предусмотрены им, но в силу общих начал и смысла гражданского законодательства порождают гражданские права и обязанности. В соответствии с этим гражданские права и обязанности возникают, изменяются и прекращаются, помимо прочих, из договоров и иных сделок, предусмотренных законодательством Республики Казахстан, а также из сделок, хотя и не предусмотренных им, но не противоречащих законодательству Республики Казахстан.",
        "Статьей 378 ГК определено, что договором признается соглашение двух или нескольких лиц об установлении, изменении или прекращении гражданских прав и обязанностей.",
        "В силу статьи 393 ГК договор считается заключенным, когда между сторонами в требуемой в подлежащих случаях форме достигнуто соглашение по всем существенным его условиям. Существенными являются условия о предмете договора, условия, которые признаны существенными законодательством или необходимы для договоров данного вида, а также все те условия, относительно которых по заявлению одной из сторон должно быть достигнуто соглашение. Если в соответствии с законодательными актами для заключения договора необходима передача имущества, договор считается заключенным с момента передачи соответствующего имущества.",
        "В соответствии со статьей 406 ГК по договору купли-продажи одна сторона (продавец) обязуется передать имущество (товар) в собственность, хозяйственное ведение или оперативное управление другой стороне (покупателю), а покупатель обязуется принять это имущество (товар) и уплатить за него определенную денежную сумму (цену).",
        f"Согласно условиям заключенного договора определено, что ответчик принял на себя обязательства по оплате стоимости полученного товара. Стоимость переданного товара с учетом скидки {format_money(discount_amount)} тенге была определена в размере {format_money(contract_total_amount)} тенге, из которых {format_money(advance_amount)} тенге были оплачены в качестве предоплаты, оставшаяся сумма в размере {format_money(installment_balance_amount)} тенге подлежала оплате в рассрочку в период с {format_date(payload.installment_from) or '—'} года по {format_date(payload.installment_to) or '—'} года равными платежами по {format_money(payload.monthly_payment_amount)} тенге.",
        f"Ответчиком были внесены платежи в размере {format_money(post_advance_payments)} тенге, в связи с чем сумма задолженности на данный момент составила {format_money(debt_amount)} тенге.",
        f"Однако обязательства, принятые на себя, ответчик на сумму в размере {format_money(debt_amount)} тенге не исполнил, что привело к нарушению прав и охраняемых законом интересов {company_name}.",
        "В силу статьи 272 ГК обязательство должно исполняться надлежащим образом в соответствии с условиями обязательства и требованиями законодательства, а при отсутствии таких условий и требований - в соответствии с обычаями делового оборота или иными обычно предъявляемыми требованиями. Согласно статье 277 ГК, если обязательство предусматривает или позволяет определить день его исполнения или период времени, в течение которого оно должно быть исполнено, обязательство подлежит исполнению в этот день или, соответственно, в любой момент в пределах такого периода.",
        "В соответствии с пунктом 1 статьи 353 Гражданского кодекса Республики Казахстан, при неисполнении или ненадлежащем исполнении денежного обязательства должник обязан уплатить кредитору неустойку (пеню) в размере 0,1% от суммы долга за каждый день просрочки.",
        f"На основании вышеуказанного положения закона, произведён расчёт пени по обязательству в размере {format_money(debt_amount)} тенге, {total_overdue_days} календарных дней просрочки, с учётом поэтапного увеличения задолженности следующим образом: {format_money(penalty_amount)} тенге, согласно нижеприведенному расчету:",
    ]
    for paragraph in main_paragraphs:
        draw_paragraph(paragraph)

    table_header_height = 76
    row_height = 44
    table_total_height = table_header_height + (len(penalty_rows) * row_height) + row_height + 42
    ensure_space(table_total_height)
    table_x = margin_left
    table_y = current_y + 8
    col_widths = [150, 150, 180, 360, 160, 210]

    def draw_cell(x: int, y: int, width: int, height: int, text: str, *, font: ImageFont.ImageFont, align: str = "center", bold: bool = False) -> None:
        draw.rectangle((x, y, x + width, y + height), outline="#111111", width=1)
        text_font = table_font_bold if bold else font
        lines = wrap_text(draw, text, text_font, max(width - 10, 20))
        bbox = draw.textbbox((0, 0), "Аг", font=text_font)
        line_height = (bbox[3] - bbox[1]) + 4
        block_height = len(lines) * line_height
        line_y = y + max((height - block_height) / 2, 4)
        for line in lines:
            line_width = draw.textlength(line, font=text_font)
            if align == "left":
                line_x = x + 6
            elif align == "right":
                line_x = x + width - line_width - 6
            else:
                line_x = x + max((width - line_width) / 2, 4)
            draw.text((line_x, line_y), line, fill="#111111", font=text_font)
            line_y += line_height

    first_header_y = table_y
    second_header_y = table_y + int(table_header_height / 2)
    x = table_x
    draw_cell(x, first_header_y, col_widths[0] + col_widths[1], int(table_header_height / 2), "Период", font=table_font, bold=True)
    x += col_widths[0] + col_widths[1]
    for label, width in zip(
        [
            "Количество дней просрочки",
            "Сумма неисполненного обязательства (тенге)",
            "Размер пени (день %)",
            "Сумма пени (тенге)",
        ],
        col_widths[2:],
    ):
        draw_cell(x, first_header_y, width, table_header_height, label, font=table_font, bold=True)
        x += width
    draw_cell(table_x, second_header_y, col_widths[0], int(table_header_height / 2), "От", font=table_font, bold=True)
    draw_cell(table_x + col_widths[0], second_header_y, col_widths[1], int(table_header_height / 2), "До", font=table_font, bold=True)

    current_row_y = table_y + table_header_height
    for row in penalty_rows:
        cell_values = [
            format_date(row["period_from"]) or "—",
            format_date(row["period_to"]) or "—",
            str(row["days"]),
            format_money(row["obligation_amount"]),
            "0,1",
            format_money(row["penalty_amount"]),
        ]
        x = table_x
        for value, width in zip(cell_values, col_widths):
            draw_cell(x, current_row_y, width, row_height, value, font=table_font)
            x += width
        current_row_y += row_height

    summary_width = sum(col_widths[:-1])
    draw_cell(table_x, current_row_y, summary_width, row_height, "Итого", font=table_font, align="right", bold=True)
    draw_cell(table_x + summary_width, current_row_y, col_widths[-1], row_height, format_money(penalty_amount), font=table_font, bold=True)
    current_y = current_row_y + row_height + 20

    closing_paragraphs = [
        f"Требования {company_name}, адресованные {client_short_name}, о необходимости исполнения обязательства оставлены последним без удовлетворения.",
        f"Совокупность приведенных норм законодательства и изложенных обстоятельств позволяет сделать вывод о том, что с {client_short_name} в пользу {company_name} подлежат взысканию сумма задолженности в размере {format_money(debt_amount)} тенге и пеня в размере {format_money(penalty_amount)} тенге.",
        "В силу статьи 4 Гражданского процессуального кодекса Республики Казахстан (далее – ГПК) задачами гражданского судопроизводства являются защита и восстановление нарушенных или оспариваемых прав, свобод и законных интересов граждан, государства и юридических лиц, соблюдение законности в гражданском обороте, обеспечение полного и своевременного рассмотрения дела, содействие мирному урегулированию спора, предупреждение правонарушений и формирование в обществе уважительного отношения к закону и суду.",
        "На основании изложенного, руководствуясь статьями 148, 149, ГПК РК",
    ]
    for paragraph in closing_paragraphs:
        draw_paragraph(paragraph)

    ensure_space(260)
    current_y = draw_text_block(
        draw,
        "Прошу:",
        x=margin_left,
        y=current_y + 6,
        width=content_width,
        font=section_font,
        line_spacing=12,
        align="left",
        paragraph_spacing=18,
    )

    ask_line = (
        f"1. Взыскать с {client_name} в пользу {company_name} сумму задолженности в размере "
        f"{format_money(debt_amount)} ({money_to_words_ru(debt_amount)}) тенге, пеню в размере "
        f"{format_money(penalty_amount)} ({money_to_words_ru(penalty_amount)}) тенге, государственную пошлину "
        f"в размере {format_money(state_duty_amount)} ({money_to_words_ru(state_duty_amount)}) тенге."
    )
    draw_paragraph(ask_line, first_line_indent=0, spacing_after=18)

    current_y = draw_text_block(
        draw,
        "Приложение:",
        x=margin_left,
        y=current_y,
        width=content_width,
        font=section_font,
        line_spacing=12,
        align="left",
        paragraph_spacing=10,
    )

    attachments = [
        "1) Квитанция об уплате государственной пошлины;",
        "2) Копия договора купли-продажи с актом приема-передачи;",
        "3) Копия устава;",
        "4) Копия свидетельства (справка о государственной регистрации);",
        "5) Копия досудебной претензии с квитанцией об отправке.",
    ]
    for item in attachments:
        draw_paragraph(item, first_line_indent=0, spacing_after=6)

    signature_y = max(current_y + 24, image_height - margin_bottom - 84)
    director_name = normalize_text(requisites.get("director_name")) or "—"
    left_signature = f"Директор {company_name}"
    draw.text((margin_left, signature_y), left_signature, fill="#111111", font=small_font)
    director_width = draw.textlength(director_name, font=small_font)
    draw.text((image_width - margin_right - director_width, signature_y), director_name, fill="#111111", font=small_font)
    generated_date = format_date(date.today()) or date.today().isoformat()
    draw.text((margin_left, signature_y + 38), generated_date, fill="#111111", font=small_font)

    pages.append(image)
    try:
        first_page, *other_pages = pages
        first_page.save(pdf_path, "PDF", resolution=150.0, save_all=True, append_images=other_pages)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF иска.") from exc

    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail="Не удалось сформировать PDF иска.")

    return pdf_path


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    user = get_optional_current_user(request)
    if user is None:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "page_title": "Legal Department Login"},
        )
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "page_title": "Legal Department",
            "app_context_json": json.dumps({"user": serialize_user(user)}, ensure_ascii=False),
        },
    )
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "page_title": "Юридический департамент"},
    )


@app.post("/api/auth/login", response_model=AuthMeResponse)
def login(payload: LoginRequest) -> Response:
    username = payload.username.strip()
    with get_connection() as connection:
        user = connection.execute(
            "SELECT * FROM users WHERE lower(username) = lower(?)",
            (username,),
        ).fetchone()
        if user is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="INVALID_CREDENTIALS")

        user_dict = dict(user)
        if not parse_bool(user_dict.get("is_active")):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="USER_DISABLED")
        if not verify_password(payload.password, str(user_dict["password_hash"]), str(user_dict["password_salt"])):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="INVALID_CREDENTIALS")

        session_token = create_session_token()
        created_at = utc_now().isoformat()
        expires_at = build_session_expiry()
        connection.execute(
            """
            INSERT INTO user_sessions (user_id, session_token, created_at, expires_at)
            VALUES (?, ?, ?, ?)
            """,
            (int(user_dict["id"]), session_token, created_at, expires_at),
        )
        connection.commit()

    response = Response(
        content=json.dumps({"user": serialize_user(user_dict)}, ensure_ascii=False),
        media_type="application/json",
    )
    set_session_cookie(response, session_token)
    return response


@app.post("/api/auth/logout")
def logout(request: Request, _user: dict[str, Any] = Depends(require_authenticated_user)) -> Response:
    session_token = request.cookies.get(SESSION_COOKIE_NAME)
    with get_connection() as connection:
        if session_token:
            connection.execute("DELETE FROM user_sessions WHERE session_token = ?", (session_token,))
            connection.commit()
    response = Response(content="{}", media_type="application/json")
    clear_session_cookie(response)
    return response


@app.get("/api/auth/me", response_model=AuthMeResponse)
def auth_me(user: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    return {"user": serialize_user(user)}


@app.post("/api/auth/change-password", response_model=AuthMeResponse)
def change_password(payload: ChangePasswordRequest, user: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    if len(payload.new_password) < PASSWORD_MIN_LENGTH:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="PASSWORD_TOO_SHORT")
    if not verify_password(payload.current_password, str(user["password_hash"]), str(user["password_salt"])):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="INVALID_CURRENT_PASSWORD")

    password_hash, password_salt = hash_password(payload.new_password)
    updated_at = utc_now().isoformat()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE users
            SET password_hash = ?, password_salt = ?, must_change_password = 0, updated_at = ?
            WHERE id = ?
            """,
            (password_hash, password_salt, updated_at, int(user["id"])),
        )
        connection.commit()
        updated_user = connection.execute("SELECT * FROM users WHERE id = ?", (int(user["id"]),)).fetchone()

    return {"user": serialize_user(updated_user)}


@app.get("/api/users", response_model=list[UserView])
def list_users(_user: dict[str, Any] = Depends(require_owner_user)) -> list[dict[str, Any]]:
    with get_connection() as connection:
        rows = connection.execute("SELECT * FROM users ORDER BY role ASC, username COLLATE NOCASE ASC").fetchall()
    return [serialize_user(row) for row in rows]


@app.post("/api/users", response_model=UserView, status_code=201)
def create_user(payload: UserCreateRequest, _user: dict[str, Any] = Depends(require_owner_user)) -> dict[str, Any]:
    role = payload.role.strip().lower()
    if role not in USER_ROLES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="INVALID_ROLE")
    if len(payload.temporary_password) < PASSWORD_MIN_LENGTH:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="PASSWORD_TOO_SHORT")

    password_hash, password_salt = hash_password(payload.temporary_password)
    timestamp = utc_now().isoformat()
    with get_connection() as connection:
        existing = connection.execute(
            "SELECT id FROM users WHERE lower(username) = lower(?)",
            (payload.username.strip(),),
        ).fetchone()
        if existing is not None:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="USERNAME_ALREADY_EXISTS")

        cursor = connection.execute(
            """
            INSERT INTO users (
                username,
                full_name,
                role,
                password_hash,
                password_salt,
                must_change_password,
                is_active,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, 1, 1, ?, ?)
            """,
            (
                payload.username.strip(),
                payload.full_name.strip(),
                role,
                password_hash,
                password_salt,
                timestamp,
                timestamp,
            ),
        )
        user_id = cursor.lastrowid
        connection.commit()
        created_user = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()

    return serialize_user(created_user)


@app.get("/api/reference-data")
def reference_data(country: str = DEFAULT_COUNTRY, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    normalized_country = normalize_country_code(country)
    with get_connection() as connection:
        catalog = build_reference_catalog(connection, normalized_country)

    return {
        "categories": CATEGORIES,
        "companies": get_companies_by_country(normalized_country),
        "cities": catalog["cities"],
        "regions": catalog["regions"],
        "courtsByCity": catalog["courtsByCity"],
        "courtsByRegion": catalog["courtsByRegion"],
        "courtCityMap": catalog["courtCityMap"],
        "courtRegionMap": catalog["courtRegionMap"],
        "cityRegionMap": catalog["cityRegionMap"],
        "citiesByRegion": catalog["citiesByRegion"],
        "decisions": DECISIONS,
        "priorityCategoryOverrides": sorted(PRIORITY_CATEGORY_OVERRIDES),
        "currentCountry": normalized_country,
        "countries": [
            {"code": code, "label": COUNTRY_LABELS[code]["ru"]}
            for code in SUPPORTED_COUNTRIES
        ],
    }


@app.get("/api/crm/debtor-prefill", response_model=CrmDebtorLookupResponse)
def crm_debtor_prefill(contract_number: str, country: str = DEFAULT_COUNTRY, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    try:
        payload = DiSellApiClient().lookup_debtor_prefill(contract_number, country=normalize_country_code(country))
    except DiSellApiError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.message) from exc

    payload["company"] = match_company_to_library(payload["company"], normalize_country_code(country))
    payload["city"] = normalize_text(payload.get("city")) or ""
    payload["address"] = normalize_text(payload.get("address"))
    payload["mobile_phone"] = normalize_text(payload.get("mobile_phone"))
    payload["home_phone"] = normalize_text(payload.get("home_phone"))
    payload["products"] = normalize_document_products(
        payload.get("products"),
        fallback_name="Товар по договору",
    )
    return payload


@app.post("/api/courts", status_code=201)
def create_court(payload: CourtCreate, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, str]:
    country = normalize_country_code(payload.country)
    name = payload.name.strip()
    city = payload.city.strip()
    region = payload.region.strip()

    with get_connection() as connection:
        catalog = build_reference_catalog(connection, country)
        if name in catalog["courtCityMap"]:
            raise HTTPException(status_code=400, detail="Суд с таким названием уже существует.")

        existing = connection.execute(
            "SELECT 1 FROM custom_courts WHERE lower(name) = lower(?) AND country = ?",
            (name, country),
        ).fetchone()
        if existing is not None:
            raise HTTPException(status_code=400, detail="Суд с таким названием уже существует.")

        created_at = datetime.now().replace(microsecond=0).isoformat()
        connection.execute(
            """
            INSERT INTO custom_courts (name, city, region, country, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (name, city, region, country, created_at),
        )
        connection.commit()

    return {"name": name, "city": city, "region": region, "country": country}


@app.get("/api/imports")
def list_import_batches(_user: dict[str, Any] = Depends(require_app_user)) -> list[dict[str, Any]]:
    with get_connection() as connection:
        batches = connection.execute(
            "SELECT * FROM import_batches ORDER BY id DESC"
        ).fetchall()
    return [serialize_import_batch(dict(batch)) for batch in batches]


@app.get("/api/imports/{batch_id}")
def get_import_batch(batch_id: int, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    with get_connection() as connection:
        batch = connection.execute(
            "SELECT * FROM import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
        if batch is None:
            raise HTTPException(status_code=404, detail="Пакет импорта не найден.")

        rows = connection.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_index ASC",
            (batch_id,),
        ).fetchall()

    return {
        "batch": serialize_import_batch(dict(batch)),
        "rows": [serialize_import_row(dict(row)) for row in rows],
    }


@app.post("/api/imports/preview-local", status_code=201)
def preview_import_file(payload: ImportPreviewRequest, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    source_path = str(Path(payload.path).expanduser())
    try:
        raw_rows = load_rows_from_path(source_path)
    except (FileNotFoundError, IsADirectoryError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    with get_connection() as connection:
        batch_id = create_import_batch(connection, source_path)
        counts = {"ok": 0, "needs_review": 0, "blocked": 0}

        for row_index, raw_row in enumerate(raw_rows, start=1):
            analyzed_row = analyze_import_row(raw_row, row_index=row_index)
            counts[analyzed_row["status"]] += 1
            connection.execute(
                """
                INSERT INTO import_rows (
                    batch_id,
                    row_index,
                    status,
                    source_data_json,
                    normalized_data_json,
                    source_category,
                    suggested_category,
                    errors_json,
                    warnings_json,
                    debtor_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    row_index,
                    analyzed_row["status"],
                    dumps_json(raw_row),
                    dumps_json(analyzed_row["normalized_data"]),
                    analyzed_row["source_category"],
                    analyzed_row["suggested_category"],
                    dumps_json(analyzed_row["errors"]),
                    dumps_json(analyzed_row["warnings"]),
                    None,
                ),
            )

        status = "needs_review" if counts["needs_review"] or counts["blocked"] else "ready"
        connection.execute(
            """
            UPDATE import_batches
            SET status = ?, total_rows = ?, ok_rows = ?, needs_review_rows = ?, blocked_rows = ?
            WHERE id = ?
            """,
            (
                status,
                len(raw_rows),
                counts["ok"],
                counts["needs_review"],
                counts["blocked"],
                batch_id,
            ),
        )
        connection.commit()

        batch = connection.execute("SELECT * FROM import_batches WHERE id = ?", (batch_id,)).fetchone()
        rows = connection.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_index ASC LIMIT 100",
            (batch_id,),
        ).fetchall()

    return {
        "batch": serialize_import_batch(dict(batch)),
        "rows": [serialize_import_row(dict(row)) for row in rows],
        "rowsTruncated": len(raw_rows) > 100,
    }


@app.post("/api/imports/safe-stage2-local", status_code=201)
def import_safe_stage2_local(payload: ImportPreviewRequest, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    source_path = str(Path(payload.path).expanduser())
    try:
        return import_safe_stage2_file(source_path)
    except (FileNotFoundError, IsADirectoryError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/imports/{batch_id}/apply")
def apply_import_batch(batch_id: int, payload: ImportApplyRequest, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    with get_connection() as connection:
        batch = connection.execute(
            "SELECT * FROM import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
        if batch is None:
            raise HTTPException(status_code=404, detail="Пакет импорта не найден.")

        if payload.import_ok_rows_only:
            import_rows = connection.execute(
                """
                SELECT * FROM import_rows
                WHERE batch_id = ? AND status = 'ok' AND debtor_id IS NULL
                ORDER BY row_index ASC
                """,
                (batch_id,),
            ).fetchall()
        else:
            import_rows = connection.execute(
                """
                SELECT * FROM import_rows
                WHERE batch_id = ? AND status != 'blocked' AND debtor_id IS NULL
                ORDER BY row_index ASC
                """,
                (batch_id,),
            ).fetchall()

        imported_count = 0
        for import_row in import_rows:
            debtor_id = insert_imported_debtor(connection, serialize_import_row(dict(import_row)))
            connection.execute(
                "UPDATE import_rows SET debtor_id = ? WHERE id = ?",
                (debtor_id, import_row["id"]),
            )
            imported_count += 1

        connection.execute(
            """
            UPDATE import_batches
            SET imported_rows = imported_rows + ?, status = ?
            WHERE id = ?
            """,
            (
                imported_count,
                "imported" if imported_count else dict(batch)["status"],
                batch_id,
            ),
        )
        connection.commit()

        updated_batch = connection.execute(
            "SELECT * FROM import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

    return {
        "batch": serialize_import_batch(dict(updated_batch)),
        "importedRows": imported_count,
    }


@app.get("/api/debtors")
def list_debtors(country: str = DEFAULT_COUNTRY, _user: dict[str, Any] = Depends(require_app_user)) -> list[dict[str, Any]]:
    normalized_country = normalize_country_code(country)
    with get_connection() as connection:
        rows = [
            dict(row)
            for row in connection.execute(
                "SELECT * FROM debtors WHERE COALESCE(country, ?) = ?",
                (DEFAULT_COUNTRY, normalized_country),
            ).fetchall()
        ]

    mark_return_rework_children(rows)
    ordered_rows = order_debtors(rows)
    return [serialize_debtor(row) for row in ordered_rows]


@app.post("/api/debtors", status_code=201)
def create_debtor(payload: DebtorCreate, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    created_at = datetime.now().replace(microsecond=0).isoformat()
    country = normalize_country_code(payload.country)

    with get_connection() as connection:
        city_value, court_value = align_city_with_court(connection, country, payload.city, payload.court)
        cursor = connection.execute(
            """
            INSERT INTO debtors (
                created_at,
                country,
                category,
                parent_debtor_id,
                client_name,
                contract_number,
                last_missed_payment_date,
                company,
                city,
                court,
                debt_amount,
                mobile_phone,
                home_phone,
                address,
                contract_total_amount,
                contract_advance_amount
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                created_at,
                country,
                DEFAULT_CATEGORY,
                None,
                payload.client_name.strip(),
                payload.contract_number.strip(),
                payload.last_missed_payment_date.isoformat(),
                payload.company.strip(),
                city_value,
                court_value,
                float(payload.debt_amount),
                normalize_text(payload.mobile_phone),
                normalize_text(payload.home_phone),
                normalize_text(payload.address),
                float(payload.contract_total_amount) if payload.contract_total_amount is not None else None,
                float(payload.contract_advance_amount) if payload.contract_advance_amount is not None else None,
            ),
        )
        debtor_id = cursor.lastrowid
        connection.commit()
        row = connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone()

    debtor = dict(row)
    debtor["has_return_rework_child"] = False
    return serialize_debtor(debtor)


@app.patch("/api/debtors/{debtor_id}")
def update_debtor(debtor_id: int, payload: DebtorUpdate, _user: dict[str, Any] = Depends(require_app_user)) -> dict[str, Any]:
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="Нет полей для обновления.")

    with get_connection() as connection:
        existing = connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="Запись не найдена.")

        existing_dict = dict(existing)
        country = normalize_country_code(str(existing_dict.get("country") or DEFAULT_COUNTRY))
        existing_dict["has_return_rework_child"] = has_return_rework_child(connection, debtor_id)
        if "last_missed_payment_date" in updates and existing_dict.get("lawsuit_installment_from"):
            updates["lawsuit_installment_from"] = updates.get("last_missed_payment_date")
        if "court" in updates:
            city_candidate = str(updates.get("city") or existing_dict.get("city") or "")
            court_candidate = str(updates.get("court") or "")
            aligned_city, aligned_court = align_city_with_court(connection, country, city_candidate, court_candidate)
            updates["city"] = aligned_city
            updates["court"] = aligned_court
        elif "city" in updates:
            updates["city"] = str(updates["city"] or "").strip()
        updates = apply_pre_validation_defaults(existing_dict, updates)
        validate_stage_dependencies(existing_dict, updates)
        updates = apply_business_rules(existing_dict, updates)
        validate_category_update(existing_dict, updates)
        validate_decision_update(existing_dict, updates)
        normalized = normalize_updates(updates)
        assignments = ", ".join(f"{field} = ?" for field in normalized)
        values = list(normalized.values()) + [debtor_id]

        connection.execute(f"UPDATE debtors SET {assignments} WHERE id = ?", values)
        row = dict(connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone())
        ensure_return_rework_child(connection, row)
        row["has_return_rework_child"] = has_return_rework_child(connection, debtor_id)
        connection.commit()

    return serialize_debtor(row)


@app.delete("/api/debtors/{debtor_id}", status_code=204, response_class=Response)
def delete_debtor(debtor_id: int, _user: dict[str, Any] = Depends(require_app_user)) -> Response:
    with get_connection() as connection:
        existing = connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="Запись не найдена.")

        row = dict(existing)
        if row.get("parent_debtor_id"):
            connection.execute("DELETE FROM debtors WHERE id = ?", (debtor_id,))
        else:
            connection.execute(
                "DELETE FROM debtors WHERE id = ? OR parent_debtor_id = ?",
                (debtor_id, debtor_id),
            )
        connection.commit()
    return Response(status_code=204)


@app.get("/api/debtors/{debtor_id}/claim-pdf")
def generate_claim_pdf(debtor_id: int, _user: dict[str, Any] = Depends(require_app_user)) -> FileResponse:
    with get_connection() as connection:
        existing = connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="Запись не найдена.")

        connection.execute(
            """
            UPDATE debtors
            SET lawsuit_installment_from = ?,
                lawsuit_installment_to = ?,
                lawsuit_monthly_payment_amount = ?,
                lawsuit_first_period_paid_amount = ?
            WHERE id = ?
            """,
            (
                payload.installment_from.isoformat(),
                payload.installment_to.isoformat(),
                float(payload.monthly_payment_amount),
                float(payload.first_period_paid_amount or 0),
                debtor_id,
            ),
        )
        row = dict(existing)
        row["has_return_rework_child"] = has_return_rework_child(connection, debtor_id)
        row["lawsuit_installment_from"] = payload.installment_from.isoformat()
        row["lawsuit_installment_to"] = payload.installment_to.isoformat()
        row["lawsuit_monthly_payment_amount"] = float(payload.monthly_payment_amount)
        row["lawsuit_first_period_paid_amount"] = float(payload.first_period_paid_amount or 0)
        connection.commit()

    pdf_path = render_claim_pdf(row)
    file_name = f"pretenziya_{debtor_id}.pdf"
    return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)


@app.post("/api/debtors/{debtor_id}/claim-pdf")
def generate_claim_pdf_with_confirmation(
    debtor_id: int,
    payload: ClaimPdfGenerateRequest,
    _user: dict[str, Any] = Depends(require_app_user),
) -> FileResponse:
    with get_connection() as connection:
        existing = connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="Запись не найдена.")

        row = dict(existing)
        row["has_return_rework_child"] = has_return_rework_child(connection, debtor_id)

    pdf_path = render_claim_pdf(
        row,
        debt_amount_override=payload.debt_amount_override,
        product_overrides=[item.model_dump() for item in (payload.product_overrides or [])],
    )
    file_name = f"pretenziya_{debtor_id}.pdf"
    return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)


@app.post("/api/debtors/{debtor_id}/lawsuit-pdf")
def generate_lawsuit_pdf_with_confirmation(
    debtor_id: int,
    payload: LawsuitPdfGenerateRequest,
    _user: dict[str, Any] = Depends(require_app_user),
) -> FileResponse:
    with get_connection() as connection:
        existing = connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="Р—Р°РїРёСЃСЊ РЅРµ РЅР°Р№РґРµРЅР°.")

        row = dict(existing)
        row["has_return_rework_child"] = has_return_rework_child(connection, debtor_id)

    pdf_path = render_lawsuit_pdf(row, payload)
    file_name = f"isk_{debtor_id}.pdf"
    return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)


def order_debtors(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    children_by_parent: dict[int, list[dict[str, Any]]] = {}
    roots: list[dict[str, Any]] = []

    for row in rows:
        parent_id = row.get("parent_debtor_id")
        if parent_id:
            children_by_parent.setdefault(int(parent_id), []).append(row)
        else:
            roots.append(row)

    roots.sort(key=debtor_sort_key, reverse=True)
    for children in children_by_parent.values():
        children.sort(key=lambda item: int(item["id"]), reverse=True)

    ordered: list[dict[str, Any]] = []

    def append_branch(row: dict[str, Any]) -> None:
        ordered.append(row)
        for child in children_by_parent.get(int(row["id"]), []):
            append_branch(child)

    for root in roots:
        append_branch(root)

    return ordered


def debtor_sort_key(row: dict[str, Any]) -> tuple[date, int]:
    contract_date = contract_date_value_from_number(str(row.get("contract_number") or ""))
    return (contract_date or date.min, int(row["id"]))


def apply_pre_validation_defaults(existing: dict[str, Any], updates: dict[str, Any]) -> dict[str, Any]:
    merged = merge_row(existing, updates)
    category = normalize_text(merged.get("category"))
    decision = normalize_text(merged.get("decision"))

    if category == CATEGORY_LAWSUIT_CLOSED and decision:
        updates["claim_sent"] = True
        updates["lawsuit_sent"] = True
        updates["lawsuit_accepted"] = True
        updates["decision_exists"] = True

    return updates


def apply_business_rules(existing: dict[str, Any], updates: dict[str, Any]) -> dict[str, Any]:
    merged = merge_row(existing, updates)
    financials = compute_financials(merged)

    if (
        existing.get("parent_debtor_id")
        and not parse_bool(existing.get("decision_exists"))
        and updates.get("decision_exists") is True
        and "category" not in updates
    ):
        updates["category"] = CATEGORY_PREPARE_LAWSUIT
        merged["category"] = CATEGORY_PREPARE_LAWSUIT

    if parse_bool(merged.get("lawsuit_sent")) and not parse_date_value(merged.get("lawsuit_sent_date")):
        updates["lawsuit_sent_date"] = date.today()
        merged["lawsuit_sent_date"] = updates["lawsuit_sent_date"]

    if parse_bool(merged.get("claim_sent")) and not parse_date_value(merged.get("claim_sent_date")):
        updates["claim_sent_date"] = date.today()
        merged["claim_sent_date"] = updates["claim_sent_date"]

    if not parse_bool(merged.get("claim_sent")):
        updates["claim_sent_date"] = None
        updates["lawsuit_sent"] = False
        updates["lawsuit_sent_date"] = None
        updates["lawsuit_accepted"] = False
        updates["hearing_date"] = None
        updates["decision_exists"] = False
        updates["decision"] = None
        updates["decision_payout"] = 0
        merged["claim_sent_date"] = None
        merged["lawsuit_sent"] = False
        merged["lawsuit_sent_date"] = None
        merged["lawsuit_accepted"] = False
        merged["hearing_date"] = None
        merged["decision_exists"] = False
        merged["decision"] = None
        merged["decision_payout"] = 0

    if not parse_bool(merged.get("lawsuit_sent")):
        updates["lawsuit_sent_date"] = None
        updates["lawsuit_accepted"] = False
        updates["hearing_date"] = None
        updates["decision_exists"] = False
        updates["decision"] = None
        updates["decision_payout"] = 0
        merged["lawsuit_sent_date"] = None
        merged["lawsuit_accepted"] = False
        merged["hearing_date"] = None
        merged["decision_exists"] = False
        merged["decision"] = None
        merged["decision_payout"] = 0

    if not parse_bool(merged.get("lawsuit_accepted")):
        updates["hearing_date"] = None
        updates["decision_exists"] = False
        updates["decision"] = None
        updates["decision_payout"] = 0
        merged["hearing_date"] = None
        merged["decision_exists"] = False
        merged["decision"] = None
        merged["decision_payout"] = 0

    if not parse_bool(merged.get("decision_exists")):
        updates["decision"] = None
        updates["decision_payout"] = 0
        merged["decision"] = None
        merged["decision_payout"] = 0

    existing_decision = normalize_text(existing.get("decision"))
    decision = normalize_text(merged.get("decision"))
    if decision != existing_decision:
        updates["decision_payout"] = 0
        merged["decision_payout"] = 0

    if decision == DECISION_SATISFY:
        updates["decision_payout"] = financials["total_amount"]
        merged["decision_payout"] = financials["total_amount"]

    return updates


def validate_category_update(existing: dict[str, Any], updates: dict[str, Any]) -> None:
    requested_category = updates.get("category")
    if requested_category is None:
        return

    normalized_requested = normalize_text(requested_category)
    if normalized_requested not in CATEGORIES:
        raise HTTPException(status_code=400, detail="Недопустимая категория.")

    if (
        not existing.get("parent_debtor_id")
        and existing.get("has_return_rework_child")
        and normalized_requested != CATEGORY_RETURNED_TO_LEGAL
    ):
        raise HTTPException(
            status_code=400,
            detail="У материнской записи с подстрокой возврата категория зафиксирована на «Возврат в работу Юр. Отдела».",
        )

    merged = merge_row(existing, updates)
    auto_category = determine_auto_category(merged)
    if (
        existing.get("parent_debtor_id")
        and not parse_bool(existing.get("decision_exists"))
        and updates.get("decision_exists") is True
        and normalized_requested == CATEGORY_PREPARE_LAWSUIT
    ):
        return
    if normalized_requested in PRIORITY_CATEGORY_OVERRIDES:
        return

    if auto_category is not None and normalized_requested != auto_category:
        raise HTTPException(
            status_code=400,
            detail="Эта категория выставляется автоматически. Доступны только приоритетные ручные статусы.",
        )

def validate_decision_update(existing: dict[str, Any], updates: dict[str, Any]) -> None:
    requested_decision = normalize_text(updates.get("decision"))
    is_return_rework = bool(existing.get("parent_debtor_id"))
    if is_return_rework and requested_decision == DECISION_RETURN:
        raise HTTPException(
            status_code=400,
            detail="Для подстроки возврата иска нельзя повторно выбрать решение «Возврат иска».",
        )

    return


def validate_stage_dependencies(existing: dict[str, Any], updates: dict[str, Any]) -> None:
    merged = merge_row(existing, updates)
    claim_sent = parse_bool(merged.get("claim_sent"))
    lawsuit_sent = parse_bool(merged.get("lawsuit_sent"))
    lawsuit_accepted = parse_bool(merged.get("lawsuit_accepted"))

    if updates.get("lawsuit_sent") is True and not claim_sent:
        raise HTTPException(
            status_code=400,
            detail="Нельзя установить «Направлен иск = Да», пока «Претензия = Нет».",
        )

    if updates.get("lawsuit_accepted") is True and (not claim_sent or not lawsuit_sent):
        raise HTTPException(
            status_code=400,
            detail="Нельзя установить «Иск принят = Да», пока претензия не отправлена и иск не направлен.",
        )

    if updates.get("decision_exists") is True and (not claim_sent or not lawsuit_sent or not lawsuit_accepted):
        raise HTTPException(
            status_code=400,
            detail="Нельзя установить «Есть решение = Да», пока претензия не отправлена, иск не направлен и иск не принят.",
        )


def ensure_return_rework_child(connection, row: dict[str, Any]) -> None:
    decision = normalize_text(row.get("decision"))
    if decision != DECISION_RETURN:
        return

    existing_child = connection.execute(
        "SELECT id FROM debtors WHERE parent_debtor_id = ? LIMIT 1",
        (row["id"],),
    ).fetchone()
    if existing_child is not None:
        return

    created_at = datetime.now().replace(microsecond=0).isoformat()
    connection.execute(
        """
        INSERT INTO debtors (
            created_at,
            country,
            category,
            parent_debtor_id,
            client_name,
            contract_number,
            last_missed_payment_date,
            company,
            city,
            court,
            claim_sent,
            claim_sent_date,
            debt_amount,
            imported_claim_sent_days,
            imported_debt_days,
            imported_penalty_amount,
            imported_state_duty_amount,
            imported_total_amount,
            lawsuit_sent,
            lawsuit_sent_date,
            lawsuit_accepted,
            hearing_date,
            decision_exists,
            decision,
            decision_payout,
            received_amount,
            comment,
            case_number
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            created_at,
            CATEGORY_PREPARE_LAWSUIT,
            row["id"],
            row["client_name"],
            row["contract_number"],
            row["last_missed_payment_date"],
            row["company"],
            row["city"],
            row["court"],
            1 if parse_bool(row.get("claim_sent")) else 1,
            row.get("claim_sent_date"),
            float(row.get("debt_amount") or 0),
            parse_int_value(row.get("imported_claim_sent_days")),
            parse_int_value(row.get("imported_debt_days")),
            float(row.get("imported_penalty_amount") or 0),
            float(row.get("imported_state_duty_amount") or 0),
            float(row.get("imported_total_amount") or 0),
            0,
            None,
            0,
            None,
            0,
            None,
            0,
            0,
            None,
            None,
        ),
    )


def has_return_rework_child(connection, debtor_id: int) -> bool:
    child = connection.execute(
        "SELECT 1 FROM debtors WHERE parent_debtor_id = ? LIMIT 1",
        (debtor_id,),
    ).fetchone()
    return child is not None


def mark_return_rework_children(rows: list[dict[str, Any]]) -> None:
    parent_ids = {int(row["parent_debtor_id"]) for row in rows if row.get("parent_debtor_id")}
    for row in rows:
        row["has_return_rework_child"] = int(row["id"]) in parent_ids


def create_import_batch(connection, source_path: str) -> int:
    created_at = datetime.now().replace(microsecond=0).isoformat()
    cursor = connection.execute(
        """
        INSERT INTO import_batches (
            created_at,
            source_path,
            source_filename,
            status,
            total_rows,
            ok_rows,
            needs_review_rows,
            blocked_rows,
            imported_rows,
            notes
        )
        VALUES (?, ?, ?, ?, 0, 0, 0, 0, 0, ?)
        """,
        (
            created_at,
            source_path,
            Path(source_path).name,
            "preview",
            "Черновой пакет импорта. Данные пока не попали в основную таблицу.",
        ),
    )
    return int(cursor.lastrowid)


def analyze_import_row(raw_row: dict[str, Any], *, row_index: int) -> dict[str, Any]:
    courts = sorted({court for values in COURTS_BY_CITY.values() for court in values})
    normalized_result = normalize_import_row(
        raw_row,
        categories=CATEGORIES,
        decisions=DECISIONS,
        companies=COMPANIES,
        cities=KAZAKHSTAN_CITIES,
        courts=courts,
    )
    normalized = normalized_result["normalized_data"]
    warnings = list(normalized_result["warnings"])
    errors = list(normalized_result["errors"])

    prepared = prepare_import_debtor_payload(normalized)
    source_category = normalize_text(normalized.get("category"))
    suggested_category = None

    if not errors:
        suggested_category = determine_auto_category(prepared) or source_category or DEFAULT_CATEGORY
        prepared["category"] = suggested_category

        if source_category and suggested_category and source_category != suggested_category:
            warnings.append(
                f"Категория из файла («{source_category}») не совпадает с категорией по правилам («{suggested_category}»)."
            )

        if normalize_text(prepared.get("decision")) and not parse_bool(prepared.get("decision_exists")):
            warnings.append("Решение указано, но признак «Есть решение» не был подтвержден в исходной строке.")

    status = "blocked" if errors else "needs_review" if warnings else "ok"
    return {
        "rowIndex": row_index,
        "status": status,
        "source_category": source_category,
        "suggested_category": suggested_category or source_category or DEFAULT_CATEGORY,
        "normalized_data": prepared,
        "warnings": dedupe_list(warnings),
        "errors": dedupe_list(errors),
    }


def build_safe_stage1_import_rows(source_path: str) -> list[dict[str, Any]]:
    courts = sorted({court for values in COURTS_BY_CITY.values() for court in values})
    safe_rows: list[dict[str, Any]] = []

    for row_index, raw_row in enumerate(load_rows_from_path(source_path), start=1):
        headers_map = build_headers_map(raw_row)
        raw_decision = normalize_text(find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["decision"]))
        if raw_decision not in SAFE_IMPORT_DECISION_CATEGORY_MAP:
            continue

        normalized_result = normalize_import_row(
            raw_row,
            categories=CATEGORIES,
            decisions=DECISIONS,
            companies=COMPANIES,
            cities=KAZAKHSTAN_CITIES,
            courts=courts,
        )
        normalized = normalized_result["normalized_data"]
        prepared = prepare_import_debtor_payload(normalized)
        prepared["company"] = prepared.get("company") or normalize_text(
            find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["company"])
        ) or ""
        prepared["city"] = prepared.get("city") or normalize_text(
            find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["city"])
        ) or ""
        prepared["court"] = prepared.get("court") or normalize_text(
            find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["court"])
        ) or ""
        prepared["category"] = SAFE_IMPORT_DECISION_CATEGORY_MAP[raw_decision]
        prepared["decision"] = None
        prepared["decision_exists"] = False
        prepared["imported_category_override"] = True
        if prepared.get("claim_sent_date") or prepared.get("imported_claim_sent_days") is not None:
            prepared["claim_sent"] = True
        if prepared.get("lawsuit_sent_date"):
            prepared["lawsuit_sent"] = True
        if prepared.get("hearing_date"):
            prepared["lawsuit_accepted"] = True

        reconstructed_date = reconstruct_last_missed_payment_date(prepared)
        prepared["last_missed_payment_date"] = reconstructed_date or ""
        prepared["client_name"] = prepared.get("client_name") or ""
        prepared["contract_number"] = prepared.get("contract_number") or ""
        prepared["company"] = prepared.get("company") or ""
        prepared["city"] = prepared.get("city") or ""
        prepared["court"] = prepared.get("court") or ""
        prepared["debt_amount"] = float(prepared.get("debt_amount") or 0)

        prepared["comment"] = normalize_text(prepared.get("comment")) or None

        safe_rows.append(
            {
                "row_index": row_index,
                "status": "ok",
                "source_data": raw_row,
                "normalized_data": prepared,
                "source_category": normalize_text(find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["category"])),
                "suggested_category": prepared["category"],
                "warnings": [],
                "errors": [],
            }
        )

    return safe_rows


def import_safe_stage1_file(source_path: str) -> dict[str, Any]:
    safe_rows = build_safe_stage1_import_rows(source_path)

    with get_connection() as connection:
        batch_id = create_import_batch(connection, source_path)
        imported_rows = 0
        skipped_rows = 0

        for safe_row in safe_rows:
            normalized = safe_row["normalized_data"]
            existing = connection.execute(
                """
                SELECT id FROM debtors
                WHERE imported_category_override = 1
                  AND client_name = ?
                  AND contract_number = ?
                  AND category = ?
                LIMIT 1
                """,
                (
                    normalized.get("client_name") or "",
                    normalized.get("contract_number") or "",
                    normalized.get("category") or DEFAULT_CATEGORY,
                ),
            ).fetchone()
            if existing is not None:
                skipped_rows += 1
                continue

            debtor_id = insert_imported_debtor(connection, safe_row)
            connection.execute(
                """
                INSERT INTO import_rows (
                    batch_id,
                    row_index,
                    status,
                    source_data_json,
                    normalized_data_json,
                    source_category,
                    suggested_category,
                    errors_json,
                    warnings_json,
                    debtor_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    safe_row["row_index"],
                    "ok",
                    dumps_json(safe_row["source_data"]),
                    dumps_json(safe_row["normalized_data"]),
                    safe_row["source_category"],
                    safe_row["suggested_category"],
                    dumps_json([]),
                    dumps_json([]),
                    debtor_id,
                ),
            )
            imported_rows += 1

        connection.execute(
            """
            UPDATE import_batches
            SET status = ?, total_rows = ?, ok_rows = ?, imported_rows = ?, notes = ?
            WHERE id = ?
            """,
            (
                "imported",
                len(safe_rows),
                len(safe_rows),
                imported_rows,
                f"Stage1 безопасный импорт статусов из решения. Пропущено дублей: {skipped_rows}.",
                batch_id,
            ),
        )
        connection.commit()

    return {
        "batch_id": batch_id,
        "total_rows": len(safe_rows),
        "imported_rows": imported_rows,
        "skipped_rows": skipped_rows,
    }


def build_safe_stage2_import_rows(source_path: str) -> list[dict[str, Any]]:
    courts = sorted({court for values in COURTS_BY_CITY.values() for court in values})
    safe_rows: list[dict[str, Any]] = []

    for row_index, raw_row in enumerate(load_rows_from_path(source_path), start=1):
        headers_map = build_headers_map(raw_row)
        raw_category = normalize_text(find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["category"])) or ""
        raw_decision = normalize_text(find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["decision"])) or ""
        if raw_decision in SAFE_IMPORT_DECISION_CATEGORY_MAP:
            continue

        rule = SAFE_IMPORT_STAGE2_RULES.get((raw_category, raw_decision))
        if rule is None:
            continue

        normalized_result = normalize_import_row(
            raw_row,
            categories=CATEGORIES,
            decisions=DECISIONS,
            companies=COMPANIES,
            cities=KAZAKHSTAN_CITIES,
            courts=courts,
        )
        normalized = normalized_result["normalized_data"]
        prepared = prepare_import_debtor_payload(normalized)
        prepared["company"] = prepared.get("company") or normalize_text(
            find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["company"])
        ) or ""
        prepared["city"] = prepared.get("city") or normalize_text(
            find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["city"])
        ) or ""
        prepared["court"] = prepared.get("court") or normalize_text(
            find_value_by_aliases(raw_row, headers_map, HEADER_ALIASES["court"])
        ) or ""
        prepared["category"] = rule["category"]
        prepared["decision"] = rule["decision"]
        prepared["decision_exists"] = rule["decision_exists"]
        prepared["claim_sent"] = rule["claim_sent"]
        prepared["lawsuit_sent"] = rule["lawsuit_sent"]
        prepared["lawsuit_accepted"] = rule["lawsuit_accepted"]
        prepared["imported_category_override"] = rule["imported_category_override"]

        reconstructed_date = reconstruct_last_missed_payment_date(prepared)
        prepared["last_missed_payment_date"] = reconstructed_date or ""
        prepared["client_name"] = prepared.get("client_name") or ""
        prepared["contract_number"] = prepared.get("contract_number") or ""
        prepared["company"] = prepared.get("company") or ""
        prepared["city"] = prepared.get("city") or ""
        prepared["court"] = prepared.get("court") or ""
        prepared["debt_amount"] = float(prepared.get("debt_amount") or 0)
        prepared["comment"] = normalize_text(prepared.get("comment")) or None

        safe_rows.append(
            {
                "row_index": row_index,
                "status": "ok",
                "source_data": raw_row,
                "normalized_data": prepared,
                "source_category": raw_category,
                "suggested_category": prepared["category"],
                "warnings": [],
                "errors": [],
            }
        )

    return safe_rows


def import_safe_stage2_file(source_path: str) -> dict[str, Any]:
    safe_rows = build_safe_stage2_import_rows(source_path)

    with get_connection() as connection:
        batch_id = create_import_batch(connection, source_path)
        imported_rows = 0
        skipped_rows = 0

        for safe_row in safe_rows:
            normalized = safe_row["normalized_data"]
            existing = connection.execute(
                """
                SELECT id FROM debtors
                WHERE client_name = ?
                  AND contract_number = ?
                  AND category = ?
                  AND COALESCE(decision, '') = COALESCE(?, '')
                LIMIT 1
                """,
                (
                    normalized.get("client_name") or "",
                    normalized.get("contract_number") or "",
                    normalized.get("category") or DEFAULT_CATEGORY,
                    normalized.get("decision"),
                ),
            ).fetchone()
            if existing is not None:
                skipped_rows += 1
                continue

            debtor_id = insert_imported_debtor(connection, safe_row)
            connection.execute(
                """
                INSERT INTO import_rows (
                    batch_id,
                    row_index,
                    status,
                    source_data_json,
                    normalized_data_json,
                    source_category,
                    suggested_category,
                    errors_json,
                    warnings_json,
                    debtor_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    safe_row["row_index"],
                    "ok",
                    dumps_json(safe_row["source_data"]),
                    dumps_json(safe_row["normalized_data"]),
                    safe_row["source_category"],
                    safe_row["suggested_category"],
                    dumps_json([]),
                    dumps_json([]),
                    debtor_id,
                ),
            )
            imported_rows += 1

        connection.execute(
            """
            UPDATE import_batches
            SET status = 'imported',
                total_rows = ?,
                ok_rows = ?,
                imported_rows = ?,
                notes = ?
            WHERE id = ?
            """,
            (
                len(safe_rows),
                len(safe_rows),
                imported_rows,
                f"Stage2 безопасный импорт процессуальных статусов и возвратов. Пропущено дублей: {skipped_rows}.",
                batch_id,
            ),
        )
        connection.commit()

    return {
        "batch_id": batch_id,
        "total_rows": len(safe_rows),
        "imported_rows": imported_rows,
        "skipped_rows": skipped_rows,
    }


def reconstruct_last_missed_payment_date(row: dict[str, Any]) -> str | None:
    claim_sent_date = parse_date_value(row.get("claim_sent_date"))
    if claim_sent_date is None:
        return None

    debt_days = parse_int_value(row.get("imported_debt_days"))
    if debt_days is not None and debt_days >= 0:
        return (claim_sent_date - timedelta(days=debt_days)).isoformat()

    debt_amount = parse_float(row.get("debt_amount"))
    if debt_amount <= 0:
        return None

    penalty_amount = row.get("imported_penalty_amount")
    if penalty_amount not in (None, ""):
        derived_days = parse_float(penalty_amount) / (debt_amount * 0.001)
        rounded_days = round(derived_days)
        if abs(derived_days - rounded_days) < 0.01 and rounded_days >= 0:
            return (claim_sent_date - timedelta(days=rounded_days)).isoformat()

    total_amount = row.get("imported_total_amount")
    if total_amount not in (None, ""):
        derived_penalty = parse_float(total_amount) - debt_amount
        if derived_penalty >= 0:
            derived_days = derived_penalty / (debt_amount * 0.001)
            rounded_days = round(derived_days)
            if abs(derived_days - rounded_days) < 0.01 and rounded_days >= 0:
                return (claim_sent_date - timedelta(days=rounded_days)).isoformat()

    return None


def prepare_import_debtor_payload(normalized: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "category": normalize_text(normalized.get("category")) or DEFAULT_CATEGORY,
        "parent_debtor_id": None,
        "client_name": normalized.get("client_name"),
        "contract_number": normalized.get("contract_number"),
        "last_missed_payment_date": normalized.get("last_missed_payment_date"),
        "company": normalized.get("company"),
        "city": normalized.get("city"),
        "court": normalized.get("court"),
        "claim_sent": normalized.get("claim_sent") if normalized.get("claim_sent") is not None else False,
        "claim_sent_date": normalized.get("claim_sent_date"),
        "debt_amount": normalized.get("debt_amount") if normalized.get("debt_amount") is not None else 0,
        "lawsuit_sent": normalized.get("lawsuit_sent") if normalized.get("lawsuit_sent") is not None else False,
        "lawsuit_sent_date": normalized.get("lawsuit_sent_date"),
        "lawsuit_accepted": (
            normalized.get("lawsuit_accepted") if normalized.get("lawsuit_accepted") is not None else False
        ),
        "hearing_date": normalized.get("hearing_date"),
        "decision_exists": (
            normalized.get("decision_exists")
            if normalized.get("decision_exists") is not None
            else bool(normalize_text(normalized.get("decision")))
        ),
        "decision": normalize_text(normalized.get("decision")),
        "decision_payout": normalized.get("decision_payout") if normalized.get("decision_payout") is not None else 0,
        "received_amount": normalized.get("received_amount") if normalized.get("received_amount") is not None else 0,
        "comment": normalize_text(normalized.get("comment")),
        "case_number": normalize_text(normalized.get("case_number")),
        "imported_claim_sent_days": parse_int_value(normalized.get("imported_claim_sent_days")),
        "imported_debt_days": parse_int_value(normalized.get("imported_debt_days")),
        "imported_penalty_amount": normalized.get("imported_penalty_amount") if normalized.get("imported_penalty_amount") is not None else 0,
        "imported_state_duty_amount": (
            normalized.get("imported_state_duty_amount") if normalized.get("imported_state_duty_amount") is not None else 0
        ),
        "imported_total_amount": normalized.get("imported_total_amount") if normalized.get("imported_total_amount") is not None else 0,
        "imported_category_override": False,
    }
    return apply_pre_validation_defaults({}, payload)


def serialize_import_batch(batch: dict[str, Any]) -> dict[str, Any]:
    created_at = datetime.fromisoformat(batch["created_at"])
    return {
        "id": batch["id"],
        "created_at": batch["created_at"],
        "created_at_label": created_at.strftime("%d.%m.%Y %H:%M"),
        "source_path": batch["source_path"],
        "source_filename": batch["source_filename"],
        "status": batch["status"],
        "total_rows": int(batch.get("total_rows") or 0),
        "ok_rows": int(batch.get("ok_rows") or 0),
        "needs_review_rows": int(batch.get("needs_review_rows") or 0),
        "blocked_rows": int(batch.get("blocked_rows") or 0),
        "imported_rows": int(batch.get("imported_rows") or 0),
        "notes": batch.get("notes"),
    }


def serialize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row["id"],
        "batch_id": row["batch_id"],
        "row_index": row["row_index"],
        "status": row["status"],
        "source_data": loads_json(row["source_data_json"]),
        "normalized_data": loads_json(row["normalized_data_json"]),
        "source_category": row.get("source_category"),
        "suggested_category": row.get("suggested_category"),
        "warnings": loads_json(row["warnings_json"]),
        "errors": loads_json(row["errors_json"]),
        "debtor_id": row.get("debtor_id"),
    }


def insert_imported_debtor(connection, import_row: dict[str, Any]) -> int:
    normalized = dict(import_row["normalized_data"])
    normalized = apply_import_insert_defaults(normalized)
    created_at = datetime.now().replace(microsecond=0).isoformat()
    cursor = connection.execute(
        """
        INSERT INTO debtors (
            created_at,
            country,
            category,
            parent_debtor_id,
            client_name,
            contract_number,
            last_missed_payment_date,
            company,
            city,
            court,
            claim_sent,
            claim_sent_date,
            debt_amount,
            imported_claim_sent_days,
            imported_debt_days,
            imported_penalty_amount,
            imported_state_duty_amount,
            imported_total_amount,
            imported_category_override,
            lawsuit_sent,
            lawsuit_sent_date,
            lawsuit_accepted,
            hearing_date,
            decision_exists,
            decision,
            decision_payout,
            received_amount,
            comment,
            case_number
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            created_at,
            normalize_country_code(str(normalized.get("country") or DEFAULT_COUNTRY)),
            normalized.get("category") or DEFAULT_CATEGORY,
            None,
            normalized["client_name"],
            normalized["contract_number"],
            normalized.get("last_missed_payment_date") or "",
            normalized["company"],
            normalized["city"],
            normalized["court"],
            1 if parse_bool(normalized.get("claim_sent")) else 0,
            normalized.get("claim_sent_date"),
            float(normalized.get("debt_amount") or 0),
            parse_int_value(normalized.get("imported_claim_sent_days")),
            parse_int_value(normalized.get("imported_debt_days")),
            float(normalized.get("imported_penalty_amount") or 0),
            float(normalized.get("imported_state_duty_amount") or 0),
            float(normalized.get("imported_total_amount") or 0),
            1 if parse_bool(normalized.get("imported_category_override")) else 0,
            1 if parse_bool(normalized.get("lawsuit_sent")) else 0,
            normalized.get("lawsuit_sent_date"),
            1 if parse_bool(normalized.get("lawsuit_accepted")) else 0,
            normalized.get("hearing_date"),
            1 if parse_bool(normalized.get("decision_exists")) else 0,
            normalized.get("decision"),
            float(normalized.get("decision_payout") or 0),
            float(normalized.get("received_amount") or 0),
            normalized.get("comment"),
            normalized.get("case_number"),
        ),
    )
    debtor_id = int(cursor.lastrowid)
    row = dict(connection.execute("SELECT * FROM debtors WHERE id = ?", (debtor_id,)).fetchone())
    ensure_return_rework_child(connection, row)
    return debtor_id


def apply_import_insert_defaults(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)

    if not parse_bool(normalized.get("decision_exists")):
        normalized["decision"] = None
        normalized["decision_payout"] = 0

    decision = normalize_text(normalized.get("decision"))
    if decision == DECISION_SATISFY and not parse_float(normalized.get("decision_payout")):
        financials = compute_financials(normalized)
        normalized["decision_payout"] = financials["total_amount"]

    return normalized


def dumps_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def loads_json(value: str | None) -> Any:
    if not value:
        return []
    return json.loads(value)


def dedupe_list(items: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def merge_row(existing: dict[str, Any], updates: dict[str, Any]) -> dict[str, Any]:
    merged = dict(existing)
    merged.update(updates)
    return merged


def normalize_updates(updates: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in updates.items():
        if isinstance(value, date):
            normalized[key] = value.isoformat()
        elif isinstance(value, bool):
            normalized[key] = int(value)
        elif value is None and key in {"last_missed_payment_date", "client_name", "contract_number", "company", "city", "court"}:
            normalized[key] = ""
        elif isinstance(value, str):
            stripped = value.strip()
            if not stripped and key in {"last_missed_payment_date", "client_name", "contract_number", "company", "city", "court"}:
                normalized[key] = ""
            else:
                normalized[key] = stripped if stripped else None
        else:
            normalized[key] = value
    return normalized


def extract_lawsuit_schedule(row: dict[str, Any]) -> dict[str, Any] | None:
    installment_from = parse_date_value(row.get("lawsuit_installment_from"))
    installment_to = parse_date_value(row.get("lawsuit_installment_to"))
    monthly_payment_amount = parse_float(row.get("lawsuit_monthly_payment_amount"))
    first_period_paid_amount = parse_float(row.get("lawsuit_first_period_paid_amount"))

    if installment_from is None or installment_to is None or monthly_payment_amount <= 0:
        return None
    if installment_to <= installment_from:
        return None

    return {
        "installment_from": installment_from,
        "installment_to": installment_to,
        "monthly_payment_amount": monthly_payment_amount,
        "first_period_paid_amount": max(first_period_paid_amount, 0.0),
    }


def serialize_debtor(row: dict[str, Any]) -> dict[str, Any]:
    created_at = datetime.fromisoformat(row["created_at"])
    today = date.today()
    last_missed_payment_date = parse_date_value(row["last_missed_payment_date"])
    claim_sent_date = parse_date_value(row.get("claim_sent_date"))
    lawsuit_sent_date = parse_date_value(row.get("lawsuit_sent_date"))
    hearing_date = parse_date_value(row.get("hearing_date"))
    birth_date = parse_date_value(row.get("birth_date"))

    financials = compute_financials(row)
    category_state = resolve_category_state(row)
    decision_exists = parse_bool(row.get("decision_exists"))
    is_hearing_overdue_without_decision = bool(
        hearing_date and hearing_date < today and not decision_exists
    )

    return {
        "id": row["id"],
        "country": normalize_country_code(str(row.get("country") or DEFAULT_COUNTRY)),
        "parent_debtor_id": row.get("parent_debtor_id"),
        "is_return_rework": bool(row.get("parent_debtor_id")),
        "entry_date": format_date(created_at.date()),
        "entry_date_iso": created_at.date().isoformat(),
        "created_at": created_at.strftime("%d.%m.%Y %H:%M"),
        "contract_date": contract_date_from_number(row["contract_number"]),
        "category": category_state["category"],
        "category_is_locked": category_state["is_locked"],
        "category_is_auto": category_state["is_auto"],
        "category_options": category_state["options"],
        "client_name": row["client_name"],
        "contract_number": row["contract_number"],
        "last_missed_payment_date": format_date(last_missed_payment_date),
        "last_missed_payment_date_iso": last_missed_payment_date.isoformat() if last_missed_payment_date else None,
        "company": sanitize_company_name(row["company"]),
        "city": row["city"],
        "court": row["court"],
        "claim_sent": parse_bool(row.get("claim_sent")),
        "claim_sent_date": format_date(claim_sent_date) if claim_sent_date else None,
        "claim_sent_date_iso": claim_sent_date.isoformat() if claim_sent_date else None,
        "claim_sent_days": financials["claim_sent_days"],
        "debt_days": financials["debt_days"],
        "debt_amount": financials["debt_amount"],
        "penalty_amount": financials["penalty_amount"],
        "state_duty_amount": financials["state_duty_amount"],
        "total_amount": financials["total_amount"],
        "lawsuit_sent": parse_bool(row.get("lawsuit_sent")),
        "lawsuit_sent_date": format_date(lawsuit_sent_date) if lawsuit_sent_date else None,
        "lawsuit_sent_date_iso": lawsuit_sent_date.isoformat() if lawsuit_sent_date else None,
        "lawsuit_accepted": parse_bool(row.get("lawsuit_accepted")),
        "hearing_date": format_date(hearing_date) if hearing_date else None,
        "hearing_date_iso": hearing_date.isoformat() if hearing_date else None,
        "decision_exists": decision_exists,
        "decision": row.get("decision"),
        "decision_payout": parse_float(row.get("decision_payout")),
        "received_amount": parse_float(row.get("received_amount")),
        "comment": row.get("comment"),
        "case_number": row.get("case_number"),
        "mobile_phone": normalize_text(row.get("mobile_phone")),
        "home_phone": normalize_text(row.get("home_phone")),
        "address": normalize_text(row.get("address")),
        "birth_date": format_date(birth_date) if birth_date else None,
        "birth_date_iso": birth_date.isoformat() if birth_date else None,
            "contract_total_amount": (
                parse_float(row.get("contract_total_amount"))
                if row.get("contract_total_amount") not in (None, "")
                else None
            ),
        "contract_advance_amount": (
            parse_float(row.get("contract_advance_amount"))
            if row.get("contract_advance_amount") not in (None, "")
            else None
        ),
        "lawsuit_installment_from": (
            parse_date_value(row.get("lawsuit_installment_from")).isoformat()
            if parse_date_value(row.get("lawsuit_installment_from"))
            else None
        ),
        "lawsuit_installment_to": (
            parse_date_value(row.get("lawsuit_installment_to")).isoformat()
            if parse_date_value(row.get("lawsuit_installment_to"))
            else None
        ),
        "lawsuit_monthly_payment_amount": (
            parse_float(row.get("lawsuit_monthly_payment_amount"))
            if row.get("lawsuit_monthly_payment_amount") not in (None, "")
            else None
        ),
        "lawsuit_first_period_paid_amount": (
            parse_float(row.get("lawsuit_first_period_paid_amount"))
            if row.get("lawsuit_first_period_paid_amount") not in (None, "")
            else None
        ),
        "case_court": row["court"],
        "is_hearing_overdue_without_decision": is_hearing_overdue_without_decision,
    }


def resolve_category_state(row: dict[str, Any]) -> dict[str, Any]:
    stored_category = normalize_text(row.get("category")) or DEFAULT_CATEGORY
    auto_category = determine_auto_category(row)
    manual_override_active = stored_category in PRIORITY_CATEGORY_OVERRIDES
    is_return_rework = bool(row.get("parent_debtor_id"))
    decision_exists = parse_bool(row.get("decision_exists"))
    has_return_rework_child = bool(row.get("has_return_rework_child"))
    imported_category_override = parse_bool(row.get("imported_category_override"))

    if not is_return_rework and has_return_rework_child:
        return {
            "category": CATEGORY_RETURNED_TO_LEGAL,
            "is_locked": True,
            "is_auto": True,
            "options": [CATEGORY_RETURNED_TO_LEGAL],
        }

    if is_return_rework and not decision_exists:
        return {
            "category": stored_category,
            "is_locked": False,
            "is_auto": False,
            "options": CATEGORIES,
        }

    if imported_category_override:
        return {
            "category": stored_category,
            "is_locked": False,
            "is_auto": False,
            "options": CATEGORIES,
        }

    if manual_override_active and auto_category is not None:
        effective_category = stored_category
        options = category_options_for(auto_category, stored_category)
        return {"category": effective_category, "is_locked": True, "is_auto": False, "options": options}

    if auto_category is not None:
        options = category_options_for(auto_category, stored_category)
        return {"category": auto_category, "is_locked": True, "is_auto": True, "options": options}

    return {
        "category": stored_category,
        "is_locked": False,
        "is_auto": False,
        "options": CATEGORIES,
    }


def category_options_for(auto_category: str | None, stored_category: str) -> list[str]:
    options: list[str] = []
    if auto_category:
        options.append(auto_category)
    if stored_category in PRIORITY_CATEGORY_OVERRIDES and stored_category not in options:
        options.append(stored_category)
    for category in CATEGORIES:
        if category in PRIORITY_CATEGORY_OVERRIDES and category not in options:
            options.append(category)
    return options


def determine_auto_category(row: dict[str, Any]) -> str | None:
    decision = normalize_text(row.get("decision"))
    lawsuit_sent = parse_bool(row.get("lawsuit_sent"))
    claim_sent = parse_bool(row.get("claim_sent"))
    financials = compute_financials(row)
    decision_payout = parse_float(row.get("decision_payout"))
    received_amount = parse_float(row.get("received_amount"))
    stored_category = normalize_text(row.get("category")) or DEFAULT_CATEGORY

    if row.get("parent_debtor_id") and not parse_bool(row.get("decision_exists")):
        return None

    if decision_payout > 0 and abs(received_amount - decision_payout) < 0.01:
        return CATEGORY_DEBT_CLOSED
    if decision == DECISION_RETURN:
        return CATEGORY_RETURNED_TO_LEGAL
    if decision == DECISION_REFUSAL:
        return CATEGORY_NO_JURISDICTION
    if decision in DECISIONS_THAT_CLOSE_LAWSUIT:
        return CATEGORY_LAWSUIT_CLOSED
    if financials["total_amount"] < 150000:
        return CATEGORY_SMALL_DEBT
    if lawsuit_sent:
        return CATEGORY_LAWSUIT_FILED
    if financials["debt_days"] is not None and financials["debt_days"] > 365 * 3:
        return CATEGORY_LIMITATION_EXPIRED
    if claim_sent and financials["claim_sent_days"] is not None and financials["claim_sent_days"] > 10:
        return CATEGORY_PREPARE_LAWSUIT
    if claim_sent:
        return CATEGORY_WAITING_FOR_CLAIM_RESPONSE
    return None


def compute_financials(row: dict[str, Any]) -> dict[str, Any]:
    today = date.today()
    last_missed_payment_date = parse_date_value(row.get("last_missed_payment_date"))
    claim_sent_date = parse_date_value(row.get("claim_sent_date"))
    debt_amount = parse_float(row.get("debt_amount"))

    if last_missed_payment_date is None:
        imported_claim_sent_days = parse_int_value(row.get("imported_claim_sent_days"))
        imported_debt_days = parse_int_value(row.get("imported_debt_days"))
        imported_penalty_amount = parse_float(row.get("imported_penalty_amount"))
        imported_total_amount = parse_float(row.get("imported_total_amount"))
        imported_state_duty_amount = parse_float(row.get("imported_state_duty_amount"))
        penalty_amount = imported_penalty_amount
        total_amount = imported_total_amount or round(debt_amount + penalty_amount, 2)
        state_duty_amount = imported_state_duty_amount or round(total_amount * 0.03, 2)
        claim_sent_days = imported_claim_sent_days if imported_claim_sent_days is not None else (
            (today - claim_sent_date).days if claim_sent_date else None
        )
        return {
            "debt_days": imported_debt_days,
            "debt_amount": debt_amount,
            "penalty_amount": penalty_amount,
            "total_amount": total_amount,
            "state_duty_amount": state_duty_amount,
            "claim_sent_days": claim_sent_days,
        }

    lawsuit_schedule = extract_lawsuit_schedule(row)
    debt_stop_date = claim_sent_date or today
    if lawsuit_schedule is not None and debt_stop_date > lawsuit_schedule["installment_from"]:
        try:
            _, penalty_amount, debt_days = build_lawsuit_penalty_rows(
                lawsuit_schedule["installment_from"],
                lawsuit_schedule["installment_to"],
                debt_stop_date,
                lawsuit_schedule["monthly_payment_amount"],
                lawsuit_schedule["first_period_paid_amount"],
            )
            total_amount = round(debt_amount + penalty_amount, 2)
            state_duty_amount = round(total_amount * 0.03, 2)
            claim_sent_days = (today - claim_sent_date).days if claim_sent_date else None
            return {
                "debt_days": debt_days,
                "debt_amount": debt_amount,
                "penalty_amount": penalty_amount,
                "total_amount": total_amount,
                "state_duty_amount": state_duty_amount,
                "claim_sent_days": claim_sent_days,
            }
        except HTTPException:
            pass

    debt_days = max((debt_stop_date - last_missed_payment_date).days, 0)
    penalty_amount = round(debt_amount * 0.001 * debt_days, 2)
    total_amount = round(debt_amount + penalty_amount, 2)
    state_duty_amount = round(total_amount * 0.03, 2)
    claim_sent_days = (today - claim_sent_date).days if claim_sent_date else None

    return {
        "debt_days": debt_days,
        "debt_amount": debt_amount,
        "penalty_amount": penalty_amount,
        "total_amount": total_amount,
        "state_duty_amount": state_duty_amount,
        "claim_sent_days": claim_sent_days,
    }


def parse_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return bool(int(value or 0))


def parse_float(value: Any) -> float:
    return round(float(value or 0), 2)


def parse_int_value(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(round(float(value)))


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def format_date(value: date | None) -> str | None:
    if value is None:
        return None
    return value.strftime("%d.%m.%Y")


def contract_date_value_from_number(contract_number: str) -> date | None:
    match = re.match(r"^(\d{2})(\d{2})(\d{2})", contract_number)
    if not match:
        return None

    day, month, year = match.groups()
    full_year = int(f"20{year}")
    try:
        return date(full_year, int(month), int(day))
    except ValueError:
        return None


def contract_date_from_number(contract_number: str) -> str | None:
    parsed = contract_date_value_from_number(contract_number)
    if parsed is None:
        return None
    return parsed.strftime("%d.%m.%Y")
